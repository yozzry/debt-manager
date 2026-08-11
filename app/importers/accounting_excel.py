"""
Specialized parser for accounting Excel exports.

Supports the multi-sheet format with:
  - 'all'    : transaction ledger (invoices/payments) — cols A:V
  - 'Data'   : customer summary list — header found by keyword scan
  - 'بيان العملاء' : reference lookups (branches, payment methods, revenue types)

Key challenges:
  - Data sheet financial columns (revenue/collected/balance) are SUMIF formulas
    that return None when data_only=True if file was never opened in Excel.
  - Code column (الكود) in Data sheet is also a formula → may be None.
  - We fall back to aggregating from the 'all' sheet transactions.
"""
import io
from datetime import datetime

from openpyxl import load_workbook


def _cell_str(cell):
    v = cell.value
    if v is None:
        return ''
    return str(v).strip()


def _cell_num(cell, default=0.0):
    v = cell.value
    if v is None:
        return default
    try:
        return float(str(v).replace(',', '').strip())
    except (ValueError, TypeError):
        return default


def _find_header_row(ws, required_words, max_scan=50):
    """Scan the first `max_scan` rows to find one that contains ALL required_words."""
    for row_idx in range(1, min(max_scan + 1, ws.max_row + 1)):
        row_vals = [_cell_str(ws.cell(row_idx, c)).lower() for c in range(1, ws.max_column + 1)]
        joined = ' '.join(row_vals)
        if all(w.lower() in joined for w in required_words):
            return row_idx
    return None


def _last_non_empty_row(ws, col=1, max_scan=5000):
    """Find the last row in `col` that is not empty, scanning upwards from max."""
    for r in range(min(ws.max_row or 0, max_scan), 0, -1):
        if ws.cell(r, col).value is not None:
            return r
    return 0


def _find_col_by_keywords(ws, header_row, keywords):
    """Find column index that matches one of the keywords in the header row."""
    for c in range(1, (ws.max_column or 20) + 1):
        val = _cell_str(ws.cell(header_row, c)).lower()
        for kw in keywords:
            if kw.lower() in val:
                return c
    return None


def detect_format(file_storage):
    """
    Detect if the uploaded file matches the accounting format.
    Returns dict with format info or None.
    """
    try:
        wb = load_workbook(file_storage, read_only=True, data_only=True)
    except Exception:
        return None

    sheet_names = [s.lower().strip() for s in wb.sheetnames]
    wb.close()

    has_all = any('all' in s for s in sheet_names)
    has_data = any('data' in s for s in sheet_names)

    if has_all and has_data:
        return {
            'type': 'accounting',
            'sheets': sheet_names,
        }
    return None


def parse_accounting_excel(file_storage):
    """
    Full parse of the accounting Excel file.
    Returns a dict with keys: transactions, customers, branches, payment_methods, revenue_types, meta.
    """
    wb = load_workbook(file_storage, data_only=True)
    result = {
        'transactions': [],
        'customers': [],
        'branches': {},
        'payment_methods': {},
        'revenue_types': {},
        'meta': {
            'filename': getattr(file_storage, 'filename', 'unknown'),
            'sheet_count': len(wb.sheetnames),
            'sheets_found': wb.sheetnames,
            'parse_time': datetime.now().isoformat(),
        },
    }

    # ── Parse 'all' sheet (transactions) ──
    ws_all = None
    for name in wb.sheetnames:
        if name.lower().strip() == 'all':
            ws_all = wb[name]
            break
    if ws_all is None and wb.sheetnames:
        ws_all = wb[wb.sheetnames[0]]

    if ws_all:
        result['transactions'] = _parse_transactions(ws_all)
        result['meta']['transaction_count'] = len(result['transactions'])

    # ── Parse 'Data' sheet (customers) ──
    ws_data = None
    for name in wb.sheetnames:
        if name.lower().strip() == 'data':
            ws_data = wb[name]
            break

    if ws_data:
        result['customers'] = _parse_customers(ws_data)
        result['meta']['customer_count'] = len(result['customers'])

    # ── Parse lookup sheet (بيان العملاء) ──
    ws_lookup = None
    for name in wb.sheetnames:
        if 'بيان' in name or 'عملاء' in name:
            ws_lookup = wb[name]
            break

    if ws_lookup:
        lookups = _parse_lookups(ws_lookup)
        result['branches'] = lookups.get('branches', {})
        result['payment_methods'] = lookups.get('payment_methods', {})
        result['revenue_types'] = lookups.get('revenue_types', {})

    wb.close()
    return result


# ─── Transaction parser ─────────────────────────────────────────────────────

TRANSACTION_HEADER_KEYWORDS = ['كود العميل', 'مدين', 'دائن']

# Column layout of the 'all' sheet in the actual accounting export:
#   col2: م | col3: كود العميل | col4: رقم التقرير | col5: المالك | col6: اسم العميل
#   col7: الموقع | col8: الصنف | col9: التاريخ | col10: الكمية | col11: السعر
#   col12: مدين | col13: دائن | col14: البيان - ملاحظات | col15: الفرع | col16: طريقة الدفع | col17: Mth
TRANSACTION_COL_FALLBACKS = {
    'seq': 2, 'customer_code': 3, 'report_num': 4, 'owner': 5, 'customer_name': 6,
    'location': 7, 'item_type': 8, 'date': 9, 'quantity': 10, 'price': 11,
    'debit': 12, 'credit': 13, 'description': 14, 'branch': 15, 'payment_method': 16, 'month': 17,
}

TRANSACTION_COL_KEYWORDS = {
    'seq': ['م', 'تسلسل'],
    'customer_code': ['كود العميل', 'كود', 'رقم الحساب', 'الحساب'],
    'report_num': ['رقم التقرير', 'تقرير'],
    'owner': ['المالك', 'مالك'],
    'customer_name': ['اسم العميل', 'اسم', 'الاسم'],
    'location': ['الموقع', 'موقع'],
    'item_type': ['الصنف', 'صنف'],
    'date': ['التاريخ', 'تاريخ'],
    'quantity': ['الكمية', 'كمية'],
    'price': ['السعر', 'سعر'],
    'debit': ['مدين', 'مديونية'],
    'credit': ['دائن', 'سداد'],
    'description': ['البيان', 'ملاحظات'],
    'branch': ['الفرع', 'فرع'],
    'payment_method': ['طريقة الدفع', 'الدفع'],
    'month': ['مth', 'month', 'الشهر'],
}


def _parse_transactions(ws):
    header_row = _find_header_row(ws, TRANSACTION_HEADER_KEYWORDS)
    if header_row is None:
        header_row = 1

    cols = {}
    for field in TRANSACTION_COL_KEYWORDS:
        found = _find_col_by_keywords(ws, header_row, TRANSACTION_COL_KEYWORDS[field])
        cols[field] = found or TRANSACTION_COL_FALLBACKS[field]

    name_col = cols['customer_name']
    last_row = _last_non_empty_row(ws, col=name_col)

    rows = []
    for r in range(header_row + 1, last_row + 1):
        code = ws.cell(r, cols['customer_code']).value
        name = ws.cell(r, name_col).value
        if code is None and name is None:
            continue

        debit = _cell_num(ws.cell(r, cols['debit']))
        credit = _cell_num(ws.cell(r, cols['credit']))
        date_val = ws.cell(r, cols['date']).value

        if isinstance(date_val, datetime):
            date_str = date_val.strftime('%Y-%m-%d')
        elif date_val is not None:
            date_str = str(date_val).strip()
        else:
            date_str = ''

        month_val = ws.cell(r, cols['month']).value
        if isinstance(month_val, datetime):
            month_str = month_val.strftime('%Y-%m') if month_val.year != 1900 else ''
        elif month_val is not None:
            month_str = str(month_val).strip()
        else:
            month_str = ''

        rows.append({
            'row': r,
            'seq': _cell_num(ws.cell(r, cols['seq'])),
            'customer_code': _cell_str(ws.cell(r, cols['customer_code'])),
            'report_num': _cell_str(ws.cell(r, cols['report_num'])),
            'owner': _cell_str(ws.cell(r, cols['owner'])),
            'customer_name': _cell_str(ws.cell(r, name_col)),
            'location': _cell_str(ws.cell(r, cols['location'])),
            'item_type': _cell_str(ws.cell(r, cols['item_type'])),
            'date': date_str,
            'quantity': _cell_num(ws.cell(r, cols['quantity'])),
            'price': _cell_num(ws.cell(r, cols['price'])),
            'debit': debit,
            'credit': credit,
            'description': _cell_str(ws.cell(r, cols['description'])),
            'branch': _cell_str(ws.cell(r, cols['branch'])),
            'payment_method': _cell_str(ws.cell(r, cols['payment_method'])),
            'month': month_str,
        })

    return rows


# ─── Customer (Data) parser ────────────────────────────────────────────────
#
# Data sheet layout (from actual file):
#   Header row:  الكود | م | كود الحساب | اسم العميل | الإيرادات | التحصيلات | الرصيد | نسبة الإيرادات % | (ref)
#   Data rows:   (code) | seq | (formula) | name | revenue | collected | balance | pct | (value)
#
# CRITICAL: All numeric columns are SUMIF formulas. With data_only=True they
# return None if the file was never opened in Excel. We must fall back to
# aggregating from the 'all' sheet.

CUSTOMER_HEADER_KEYWORDS = ['الكود', 'اسم العميل', 'الإيرادات', 'التحصيلات', 'الرصيد']


def _parse_customers(ws):
    header_row = _find_header_row(ws, CUSTOMER_HEADER_KEYWORDS)
    if header_row is None:
        header_row = 1

    code_col = _find_col_by_keywords(ws, header_row, ['الكود']) or 1
    seq_col = _find_col_by_keywords(ws, header_row, ['م', 'تسلسل', '序号']) or 2
    acct_col = _find_col_by_keywords(ws, header_row, ['كود الحساب', 'حساب']) or 3
    name_col = _find_col_by_keywords(ws, header_row, ['اسم العميل', 'اسم', 'الاسم']) or 4
    rev_col = _find_col_by_keywords(ws, header_row, ['الإيرادات', 'إيرادات', 'الايرادات', 'ايرادات', 'الإيراد']) or 5
    coll_col = _find_col_by_keywords(ws, header_row, ['التحصيلات', 'تحصيلات', 'التحصيل', 'التحصيلات']) or 6
    bal_col = _find_col_by_keywords(ws, header_row, ['الرصيد', 'رصيد']) or 7
    pct_col = _find_col_by_keywords(ws, header_row, ['نسبة', '%']) or 8
    ref_col = _find_col_by_keywords(ws, header_row, ['مرجع', 'ref'])

    last_row = _last_non_empty_row(ws, col=name_col)

    rows = []
    for r in range(header_row + 1, last_row + 1):
        name = ws.cell(r, name_col).value
        if name is None or str(name).strip() == '':
            continue

        code_val = ws.cell(r, code_col).value
        revenue = _cell_num(ws.cell(r, rev_col))
        collected = _cell_num(ws.cell(r, coll_col))
        balance = _cell_num(ws.cell(r, bal_col))

        rows.append({
            'row': r,
            'code': _cell_str(ws.cell(r, code_col)) if code_val is not None else '',
            'seq': _cell_num(ws.cell(r, seq_col)),
            'account_code': _cell_str(ws.cell(r, acct_col)),
            'name': str(name).strip(),
            'revenue': revenue,
            'collected': collected,
            'balance': balance,
            'revenue_pct': _cell_num(ws.cell(r, pct_col)),
            'ref_code': _cell_str(ws.cell(r, ref_col)) if ref_col else '',
        })

    return rows


# ─── Lookup (بيان العملاء) parser ──────────────────────────────────────────
#
# The lookup sheet's header row contains the section labels spread across columns:
#   'الفرع'      → branch data   in that column (code in the adjacent column)
#   'طريقة الدفع' → payment data  in that column (code in the adjacent column)
#   'نوع الايراد' → revenue data  in that column (code in the adjacent column)

def _parse_lookups(ws):
    branches = {}
    payment_methods = {}
    revenue_types = {}

    header_row = _find_header_row(ws, ['الفرع', 'طريقة الدفع', 'نوع الايراد'])
    if header_row is None:
        header_row = _find_header_row(ws, ['الفرع', 'طريقة الدفع'])
    if header_row is None:
        return {
            'branches': branches,
            'payment_methods': payment_methods,
            'revenue_types': revenue_types,
        }

    branch_col = _find_col_by_keywords(ws, header_row, ['الفرع'])
    pay_col = _find_col_by_keywords(ws, header_row, ['طريقة الدفع'])
    rev_col = _find_col_by_keywords(ws, header_row, ['نوع الايراد', 'نوع الإيراد', 'الايراد', 'الإيراد'])

    def _read_pair(label_col):
        pairs = {}
        if not label_col:
            return pairs
        for rr in range(header_row + 1, (ws.max_row or 0) + 1):
            name = ws.cell(rr, label_col).value
            code = ws.cell(rr, label_col - 1).value
            if code is None:
                code = ws.cell(rr, label_col + 1).value
            if name is None and code is None:
                break
            if name is None or str(name).strip() == '':
                continue
            pairs[str(code).strip() if code is not None else ''] = str(name).strip()
        return pairs

    if branch_col:
        branches = _read_pair(branch_col)
    if pay_col:
        payment_methods = _read_pair(pay_col)
    if rev_col:
        revenue_types = _read_pair(rev_col)

    return {
        'branches': branches,
        'payment_methods': payment_methods,
        'revenue_types': revenue_types,
    }


# ─── Helpers for import route ──────────────────────────────────────────────

def build_customer_preview(parsed):
    """
    Merge transaction data with customer summary for a rich preview.
    When Data sheet formula values are None (file not opened in Excel),
    we compute revenue/collected/balance from the 'all' sheet transactions.
    """
    customers = parsed['customers']
    transactions = parsed['transactions']

    tx_by_code = {}
    for tx in transactions:
        code = tx['customer_code']
        if not code:
            continue
        if code not in tx_by_code:
            tx_by_code[code] = {'total_debit': 0, 'total_credit': 0, 'tx_count': 0}
        tx_by_code[code]['total_debit'] += tx['debit']
        tx_by_code[code]['total_credit'] += tx['credit']
        tx_by_code[code]['tx_count'] += 1

    tx_by_name = {}
    for tx in transactions:
        name = tx['customer_name']
        if not name:
            continue
        if name not in tx_by_name:
            tx_by_name[name] = {'total_debit': 0, 'total_credit': 0, 'tx_count': 0}
        tx_by_name[name]['total_debit'] += tx['debit']
        tx_by_name[name]['total_credit'] += tx['credit']
        tx_by_name[name]['tx_count'] += 1

    preview = []
    for c in customers:
        code = c['code']
        name = c['name']
        tx_stats = tx_by_code.get(code) or tx_by_name.get(name, {'total_debit': 0, 'total_credit': 0, 'tx_count': 0})

        rev = c.get('revenue', 0) or 0
        coll = c.get('collected', 0) or 0
        bal = c.get('balance', 0)

        if rev == 0 and tx_stats['total_debit'] > 0:
            rev = tx_stats['total_debit']
        if coll == 0 and tx_stats['total_credit'] > 0:
            coll = tx_stats['total_credit']
        if bal == 0:
            bal = rev - coll

        preview.append({
            'code': code,
            'name': name,
            'revenue': rev,
            'collected': coll,
            'balance': bal,
            'tx_count': tx_stats['tx_count'],
            'account_code': c.get('account_code', ''),
            'ref_code': c.get('ref_code', ''),
        })

    return preview
