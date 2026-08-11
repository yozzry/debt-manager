import os
import sys
import re
import uuid
import random
import time
import logging
import shutil
import subprocess
import threading
from logging.handlers import RotatingFileHandler
from datetime import datetime, date, timedelta

_baileys_setup_lock = threading.Lock()

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp', 'pdf'}

_background_logger = None


def _get_background_logger():
    global _background_logger
    if _background_logger is not None:
        return _background_logger
    try:
        from flask import current_app
        base_dir = current_app.config.get('BASE_DIR')
        if base_dir is None:
            raise RuntimeError('BASE_DIR not set')
    except (RuntimeError, KeyError):
        if getattr(sys, 'frozen', False):
            base_dir = os.path.dirname(sys.executable)
        else:
            base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    logs_dir = os.path.join(base_dir, 'logs')
    os.makedirs(logs_dir, exist_ok=True)
    _background_logger = logging.getLogger('background_tasks')
    _background_logger.setLevel(logging.ERROR)
    if not _background_logger.handlers:
        handler = RotatingFileHandler(
            os.path.join(logs_dir, 'background_errors.log'),
            maxBytes=2 * 1024 * 1024, backupCount=3, encoding='utf-8'
        )
        handler.setFormatter(logging.Formatter('%(asctime)s [%(levelname)s] %(message)s'))
        _background_logger.addHandler(handler)
    return _background_logger


def log_background_error(message, exc=None):
    logger = _get_background_logger()
    extra = f"\n{type(exc).__name__}: {exc}" if exc else ""
    logger.error(f"{message}{extra}")


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def landing_url():
    """الصفحة الافتراضية حسب وضع التشغيل (debt = المديونية / commerce = التجارة)."""
    from flask import session, url_for
    mode = session.get('app_mode', 'debt')
    if mode == 'commerce':
        return url_for('dashboard.index')
    return url_for('clients.index')


def recalc_client(client_id):
    from app.models import Client, db
    c = db.session.get(Client, client_id)
    if not c:
        return
    c.total_debt = (float(c.base_debt) if c.base_debt else 0.0) + float(sum(i.amount for i in c.invoices))
    c.total_paid = (float(c.base_paid) if c.base_paid else 0.0) + float(sum(p.amount for p in c.payments))
    c.status = 'paid' if c.balance <= 0 else 'due'
    db.session.commit()


def log_activity(user_id, action, entity_type=None, entity_id=None, details=None, ip_address=None):
    from app.models import ActivityLog, db
    entry = ActivityLog(
        user_id=user_id,
        action=action,
        entity_type=entity_type,
        entity_id=entity_id,
        details=details,
        ip_address=ip_address,
    )
    db.session.add(entry)
    try:
        db.session.flush()
    except Exception:
        db.session.rollback()


def update_stock(product, quantity, movement_type='ADJUST', reference=None, notes=None, user_id=None):
    """تعديل رصيد منتج مع تسجيل حركة مخزون.
    movement_type: IN (زيادة) / OUT (نقص) / ADJUST (تحديد رصيد مباشر).
    تُرجع (ok, message) — وتعمل فقط داخل جلسة قاعدة بيانات نشطة."""
    from app.models import StockMovement, db
    if movement_type == 'OUT':
        new_qty = float(product.current_stock or 0) - float(quantity)
    elif movement_type == 'IN':
        new_qty = float(product.current_stock or 0) + float(quantity)
    else:
        new_qty = float(quantity)
    if new_qty < 0:
        return False, f'الكمية المطلوبة أكبر من المتاح (المتوفر: {product.current_stock:g})'
    product.current_stock = new_qty
    db.session.add(StockMovement(
        product_id=product.id,
        movement_type=movement_type,
        quantity=quantity,
        balance_after=new_qty,
        reference=reference,
        notes=notes,
        created_by=user_id,
    ))
    try:
        db.session.flush()
    except Exception:
        db.session.rollback()
        return False, 'فشل حفظ حركة المخزون'
    return True, 'تم تحديث المخزون'


def get_low_stock_products():
    """المنتجات النشطة التي نفد مخزونها أو انخفض عن الحد الأدنى."""
    from app.models import Product
    return [p for p in Product.query.filter_by(is_active=True).all()
            if p.stock_status != 'ok']


COUNTRY_OPTIONS = {
    'EG': {'name': 'مصر', 'phone_code': '20', 'timezone': 'Africa/Cairo', 'currency': 'جنيه مصري', 'currency_short': 'ج.م'},
    'SA': {'name': 'السعودية', 'phone_code': '966', 'timezone': 'Asia/Riyadh', 'currency': 'ريال سعودي', 'currency_short': 'ر.س'},
}


def get_app_settings():
    from app.models import Settings
    country = Settings.get('app_country', 'EG')
    info = COUNTRY_OPTIONS.get(country, COUNTRY_OPTIONS['EG'])
    return {
        'country': country,
        'country_name': info['name'],
        'phone_code': info['phone_code'],
        'timezone': Settings.get('app_timezone', info['timezone']),
        'currency': Settings.get('app_currency', info['currency']),
        'currency_short': Settings.get('app_currency_short', info['currency_short']),
        'company_name': Settings.get('company_name', ''),
    }


def get_company_name():
    """اسم المنشأة/الشركة للإظهار على التقارير المالية (قد يكون فارغاً)."""
    from app.models import Settings
    return (Settings.get('company_name', '') or '').strip()


def get_currency_short():
    """اختصار العملة المعروض على التقارير المالية (مثل ج.م / ر.س)."""
    from app.models import Settings
    return (Settings.get('app_currency_short', '') or '').strip()


def currency_suffix():
    """لاحقة العملة لرؤوس أعمدة المبالغ مثل: المبلغ (ج.م)."""
    cur = get_currency_short()
    return f' ({cur})' if cur else ''


def company_header_elements(font_name='Arabic'):
    """عنصر ترويسة اسم المنشأة للتقارير المالية (أو [] عند عدم وجود اسم)."""
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import Paragraph
    name = get_company_name()
    if not name:
        return []
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle('ArCompany', parent=styles['Normal'],
                              fontName=font_name, fontSize=13, spaceAfter=2,
                              alignment=1, textColor=colors.HexColor('#111827')))
    return [Paragraph(name, styles['ArCompany'])]


def get_whatsapp_settings():
    from app.models import Settings
    return {
        'baileys_url': Settings.get('baileys_url', 'http://localhost:3001'),
    }


def normalize_phone(phone, country='EG'):
    """Convert a local phone number to WhatsApp international format.

    Prefix detection takes priority over the app country, so Saudi mobiles
    (05...) convert to 9665... even when the app country is Egypt, and
    Egyptian mobiles (01x...) convert to 201x... even when it is Saudi:
      SA mobile:  05xxxxxxxx  -> 9665xxxxxxxx
      EG mobile:  01x...      -> 201x...
      EG landline: 02xxxxxxx  -> 202xxxxxxx
    Numbers already in international form (966..., 20...) or prefixed with a
    country code are kept as-is. A leading "00" is stripped first.
    """
    p = re.sub(r'\D', '', str(phone or ''))
    if not p:
        return phone or ''
    if p.startswith('00'):
        p = p[2:]
    if p.startswith('0'):
        if p.startswith('05'):
            return '966' + p[1:]
        if p.startswith('01') or p.startswith('02'):
            return '20' + p[1:]
        code = COUNTRY_OPTIONS.get(country, COUNTRY_OPTIONS['EG'])['phone_code']
        return code + p[1:]
    return p


def send_whatsapp(phone, message):
    from app.models import Settings
    phone = normalize_phone(phone, Settings.get('app_country', 'EG'))
    ws = get_whatsapp_settings()
    base = ws['baileys_url']
    try:
        r = requests.get(f"{base}/status", timeout=5)
        data = r.json()
        if data.get('status') != 'connected':
            return False, 'WhatsApp غير متصل — افتح صفحة الإعدادات وامسح QR ثم أعد المسح'
    except requests.ConnectionError:
        return False, 'خدمة Baileys غير شغالة — شغّلها من صفحة الإعدادات أولاً'
    except requests.Timeout:
        return False, 'خدمة Baileys لا تستجيب — تحقق من المنفذ 3001'
    except Exception as e:
        from flask import current_app
        current_app.logger.warning(f"WhatsApp status check failed: {e}")
        return False, 'فشل التحقق من حالة Baileys'
    try:
        url = f"{base}/send"
        r = requests.post(url, json={'to': phone, 'message': message}, timeout=60)
        data = r.json()
        ok = data.get('success', False)
        return ok, data.get('error', r.text)
    except requests.ConnectionError:
        return False, 'خدمة Baileys غير شغالة — شغّلها من صفحة الإعدادات أولاً'
    except requests.Timeout:
        return False, 'انتهت مهلة الإرسال — تحقق من اتصال واتساب'
    except Exception as e:
        from flask import current_app
        current_app.logger.error(f"WhatsApp send failed: {e}")
        return False, 'حدث خطأ أثناء الإرسال'


def build_reminder_message(client, template_num=1):
    from app.models import Settings
    app = get_app_settings()
    cur = app['currency_short']
    t1 = Settings.get('template_1',
        f'السلام عليكم {{name}}، تذكير بأن لديك رصيد مستحق بقيمة {{balance}} {cur}. نرجو السداد في أقرب وقت.')
    t2 = Settings.get('template_2',
        f'عزيزي/عزيزتي {{name}}، يُرجى العلم بأن مديونيتك المستحقة بلغت {{balance}} {cur}.')
    t3 = Settings.get('template_3',
        f'{{name}}، رصيدك المستحق: {{balance}} {cur}. للاستفسار تواصل معنا.')
    templates = {1: t1, 2: t2, 3: t3}
    link = Settings.get('payment_link', '')
    msg = templates.get(template_num, t1)
    msg = msg.replace('{name}', client.name).replace('{balance}', f"{client.balance:,.2f}")
    if link:
        msg += f"\n{link}"
    return msg


def ensure_baileys_ready(baileys_dir=None):
    """تجهيز خدمة واتساب تلقائيًا: تثبيت مكتبات node إن كانت غائبة (مثل
    install_baileys.bat لكن بدون أي خطوة يدوية) والتحقق من جاهزية التشغيل.
    آمنة للتكرار (قفل يمنع تشغيل npm مرتين). تُرجع (ok, message)."""
    if baileys_dir is None:
        try:
            from flask import current_app
            baileys_dir = os.path.join(
                current_app.config.get('BASE_DIR') or os.path.dirname(current_app.instance_path),
                'baileys_service')
        except Exception:
            return False, 'تعذر تحديد مجلد خدمة واتساب'

    index_js = os.path.join(baileys_dir, 'index.js')
    if not os.path.isfile(index_js):
        return False, 'ملف index.js غير موجود في baileys_service'

    node = shutil.which('node')
    if not node:
        return False, 'Node.js غير مثبت — ثبّته من https://nodejs.org ثم أعد تشغيل البرنامج'

    npm = shutil.which('npm') or shutil.which('npm.cmd')
    if not npm:
        return False, 'npm غير موجود في PATH — أعد تثبيت Node.js ثم أعد تشغيل البرنامج'

    dep_dir = os.path.join(baileys_dir, 'node_modules', '@whiskeysockets', 'baileys')
    with _baileys_setup_lock:
        if not os.path.isdir(dep_dir):
            try:
                from flask import current_app
                current_app.logger.info('baileys: node_modules missing, running npm install automatically')
            except Exception:
                pass
            git = shutil.which('git')
            if git:
                try:
                    subprocess.run(
                        ['git', 'config', '--global', 'url."https://github.com/".insteadOf', 'ssh://git@github.com/'],
                        capture_output=True, timeout=30, creationflags=subprocess.CREATE_NO_WINDOW)
                    subprocess.run(
                        ['git', 'config', '--global', 'url."https://github.com/".insteadOf', 'git@github.com:'],
                        capture_output=True, timeout=30, creationflags=subprocess.CREATE_NO_WINDOW)
                except Exception:
                    pass
            try:
                result = subprocess.run(
                    [npm, 'install', '--no-audit', '--no-fund'],
                    cwd=baileys_dir, capture_output=True, text=True, timeout=900,
                    creationflags=subprocess.CREATE_NO_WINDOW)
            except Exception as e:
                return False, f'فشل تشغيل npm install: {e}'
            if not os.path.isdir(dep_dir):
                tail = (result.stderr or result.stdout or '')[-300:]
                return False, f'فشل تثبيت مكتبات واتساب (تحقق من الإنترنت ثم أعد تشغيل البرنامج). التفاصيل: {tail}'
    if not os.path.isdir(dep_dir):
        return False, 'مكتبات واتساب غير مثبتة — أعد تشغيل البرنامج لتثبيتها تلقائيًا'
    return True, 'الخدمة جاهزة'


def start_baileys_bridge(baileys_dir=None):
    """تشغيل جسر واتساب (node index.js) بشكل منفصل إن لم يكن يعمل بالفعل على
    المنفذ 3001. تُرجع (ok, message)."""
    import socket
    if baileys_dir is None:
        try:
            from flask import current_app
            baileys_dir = os.path.join(
                current_app.config.get('BASE_DIR') or os.path.dirname(current_app.instance_path),
                'baileys_service')
        except Exception:
            return False, 'تعذر تحديد مجلد خدمة واتساب'
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        s.settimeout(1)
        already = s.connect_ex(('127.0.0.1', 3001)) == 0
        s.close()
        if already:
            return True, 'الخدمة تعمل بالفعل'
    except Exception:
        pass
    node = shutil.which('node')
    if not node:
        return False, 'Node.js غير مثبت'
    log_path = os.path.join(baileys_dir, 'baileys.log')
    try:
        log_file = open(log_path, 'a', encoding='utf-8')
    except Exception:
        log_file = open(os.path.join(baileys_dir, 'baileys_setup.log'), 'a', encoding='utf-8')
    try:
        subprocess.Popen(
            ['node', 'index.js'],
            cwd=baileys_dir,
            stdout=log_file,
            stderr=log_file,
            creationflags=subprocess.DETACHED_PROCESS | subprocess.CREATE_NEW_PROCESS_GROUP,
            close_fds=True,
        )
    finally:
        log_file.close()
    return True, 'تم تشغيل خدمة واتساب — امسح رمز QR من صفحة الإعدادات'


def _send_reminders_background(app, trigger_time=None, trigger_freq=None, trigger_day=None, trigger_dom=None):
    with app.app_context():
        try:
            from app.models import Client
            from app.database import Settings
            app.logger.info(f"Running scheduled reminders (time={trigger_time} freq={trigger_freq})...")
            due_clients = Client.query.filter_by(status='due', reminder_enabled=True).all()
            global_freq = Settings.get('reminder_frequency', 'daily')
            global_day = Settings.get('reminder_day', 'sun')
            global_dom = int(Settings.get('reminder_dom', '1'))
            global_times_str = Settings.get('reminder_times', '10:00')

            for c in due_clients:
                try:
                    if not c.phone:
                        continue
                    freq = c.reminder_frequency or global_freq
                    day = c.reminder_day or global_day
                    dom = c.reminder_dom if c.reminder_dom is not None else global_dom
                    if trigger_freq != freq:
                        continue
                    if freq == 'weekly' and trigger_day != day:
                        continue
                    if freq == 'monthly' and str(trigger_dom) != str(dom):
                        continue
                    effective_times = c.reminder_times or global_times_str
                    if trigger_time not in [t.strip() for t in effective_times.split(',') if t.strip()]:
                        continue
                    msg = build_reminder_message(c, c.reminder_template or 1)
                    delay = random.uniform(3, 8)
                    time.sleep(delay)
                    ok, resp = send_whatsapp(c.phone, msg)
                    app.logger.info(f"Reminder to {c.name} ({c.phone}): {'OK' if ok else 'FAIL'} - {resp}")
                except Exception as e:
                    log_background_error(
                        f"Failed to send reminder to client #{c.id} ({c.name}): {e}", e
                    )
                    continue
        except Exception as e:
            log_background_error(f"Background reminders task failed entirely: {e}", e)
            app.logger.error(f"Background reminders failed: {e}")


def send_scheduled_reminders(app, trigger_time=None, trigger_freq=None, trigger_day=None, trigger_dom=None):
    import threading
    t = threading.Thread(target=_send_reminders_background,
                         args=(app, trigger_time, trigger_freq, trigger_day, trigger_dom), daemon=True)
    t.start()


def create_pdf_report(clients, title="تقرير المديونيات", report_type="full"):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm, mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import io

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=2*cm, bottomMargin=2*cm)

    try:
        pdfmetrics.registerFont(TTFont('Arabic', 'C:/Windows/Fonts/arial.ttf'))
        font_name = 'Arabic'
    except Exception:
        font_name = 'Helvetica'

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle('ArTitle', parent=styles['Title'],
                              fontName=font_name, fontSize=18, spaceAfter=20,
                              alignment=1))
    styles.add(ParagraphStyle('ArSubtitle', parent=styles['Normal'],
                              fontName=font_name, fontSize=11, spaceAfter=10,
                              alignment=1, textColor=colors.grey))

    elements = []
    elements.append(Paragraph(title, styles['ArTitle']))
    elements.append(Paragraph(f"تاريخ التقرير: {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles['ArSubtitle']))
    elements.append(Spacer(1, 0.5*cm))

    header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
    header_font = Font(name=font_name, bold=True, color='FFFFFF', size=10)

    headers = ['#', 'اسم العميل', 'الهاتف', 'المديونية', 'المدفوع', 'المتبقي', 'الحالة']
    col_widths = [1.2*cm, 5*cm, 3.5*cm, 3*cm, 3*cm, 3*cm, 2.5*cm]

    data = [headers]
    total_debt = 0
    total_paid = 0
    total_balance = 0

    for i, c in enumerate(clients, 1):
        total_debt += c.total_debt
        total_paid += c.total_paid
        total_balance += c.balance
        status_ar = 'مدفوع' if c.status == 'paid' else 'مستحق'
        data.append([
            str(i), c.name, c.phone or '—',
            f"{c.total_debt:,.2f}", f"{c.total_paid:,.2f}",
            f"{c.balance:,.2f}", status_ar
        ])

    data.append(['', 'الإجمالي', '',
                 f"{total_debt:,.2f}", f"{total_paid:,.2f}",
                 f"{total_balance:,.2f}", ''])

    table_style = TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), font_name),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('ALIGN', (3, 1), (5, -1), 'CENTER'),
        ('ALIGN', (6, 1), (6, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F8FAFC')]),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#F1F5F9')),
        ('FONTNAME', (0, -1), (-1, -1), font_name),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ])

    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(table_style)
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)
    return buffer


def create_period_comparison(clients_current, clients_previous,
                             current_label="الشهر الحالي", previous_label="الشهر الماضي"):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import io

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=2*cm, bottomMargin=2*cm)

    try:
        pdfmetrics.registerFont(TTFont('Arabic', 'C:/Windows/Fonts/arial.ttf'))
        font_name = 'Arabic'
    except Exception:
        font_name = 'Helvetica'

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle('ArTitle', parent=styles['Title'],
                              fontName=font_name, fontSize=16, spaceAfter=15, alignment=1))

    elements = []
    elements.append(Paragraph("مقارنة بين الفترات", styles['ArTitle']))
    elements.append(Spacer(1, 0.5*cm))

    cur_debt = sum(c.total_debt for c in clients_current)
    cur_paid = sum(c.total_paid for c in clients_current)
    prev_debt = sum(c.total_debt for c in clients_previous)
    prev_paid = sum(c.total_paid for c in clients_previous)

    data = [
        ['المقياس', current_label, previous_label, 'الفرق'],
        ['عدد العملاء', str(len(clients_current)), str(len(clients_previous)),
         str(len(clients_current) - len(clients_previous))],
        ['إجمالي المديونية', f"{cur_debt:,.2f}", f"{prev_debt:,.2f}",
         f"{cur_debt - prev_debt:+,.2f}"],
        ['إجمالي المدفوع', f"{cur_paid:,.2f}", f"{prev_paid:,.2f}",
         f"{cur_paid - prev_paid:+,.2f}"],
        ['الرصيد المتبقي', f"{cur_debt - cur_paid:,.2f}", f"{prev_debt - prev_paid:,.2f}",
         f"{(cur_debt - cur_paid) - (prev_debt - prev_paid):+,.2f}"],
    ]

    table = Table(data, colWidths=[5*cm, 5*cm, 5*cm, 5*cm])
    table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), font_name),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
        ('BACKGROUND', (0, 1), (0, -1), colors.HexColor('#F1F5F9')),
        ('ROWBACKGROUNDS', (1, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)
    return buffer


def create_ledger_pdf(sections, title="دفتر الأستاذ", account_label="",
                      date_from="", date_to=""):
    """PDF لدفتر الأستاذ: قسم لكل حساب مع الرصيد الجاري (كشف حساب)."""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import io

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            leftMargin=1.2*cm, rightMargin=1.2*cm,
                            topMargin=1.8*cm, bottomMargin=1.8*cm)

    try:
        pdfmetrics.registerFont(TTFont('Arabic', 'C:/Windows/Fonts/arial.ttf'))
        font_name = 'Arabic'
    except Exception:
        font_name = 'Helvetica'

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle('ArTitle', parent=styles['Title'],
                              fontName=font_name, fontSize=17, spaceAfter=8,
                              alignment=1))
    styles.add(ParagraphStyle('ArSubtitle', parent=styles['Normal'],
                              fontName=font_name, fontSize=10, spaceAfter=8,
                              alignment=1, textColor=colors.grey))
    styles.add(ParagraphStyle('ArSection', parent=styles['Normal'],
                              fontName=font_name, fontSize=11, spaceAfter=6,
                              textColor=colors.HexColor('#4F46E5')))

    elements = []
    elements += company_header_elements(font_name)
    elements.append(Paragraph(title, styles['ArTitle']))
    subtitle = f"تاريخ التقرير: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    if account_label:
        subtitle += f" — الحساب: {account_label}"
    if date_from or date_to:
        subtitle += f" — الفترة: {date_from or '...'} إلى {date_to or '...'}"
    elements.append(Paragraph(subtitle, styles['ArSubtitle']))
    elements.append(Spacer(1, 0.4*cm))

    headers = ['#', 'القيد', 'التاريخ', 'البيان', f'مدين{currency_suffix()}',
               f'دائن{currency_suffix()}', f'الرصيد الجاري{currency_suffix()}']
    col_widths = [1*cm, 3.2*cm, 2.2*cm, 6.5*cm, 2.6*cm, 2.6*cm, 3*cm]

    table_style = TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), font_name),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('ALIGN', (4, 1), (6, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ])

    for section in sections:
        account = section['account']
        section_label = (f"{account.code} - {account.name} ({account.type_label}) "
                         f"— الرصيد الافتتاحي: {section['opening']:,.2f}")
        elements.append(Paragraph(section_label, styles['ArSection']))
        data = [headers]
        for i, row in enumerate(section['rows'], 1):
            data.append([
                str(i),
                row['entry_number'],
                row['date'].strftime('%Y-%m-%d') if row['date'] else '',
                row['description'] or '',
                f"{row['debit']:,.2f}" if row['debit'] else '',
                f"{row['credit']:,.2f}" if row['credit'] else '',
                f"{row['running']:,.2f}",
            ])
        data.append(['', 'الرصيد الختامي', '', '',
                     '', '', f"{section['closing']:,.2f}"])
        table = Table(data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), font_name),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('ALIGN', (4, 1), (6, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2),
             [colors.white, colors.HexColor('#F8FAFC')]),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#F1F5F9')),
            ('FONTNAME', (0, -1), (-1, -1), font_name),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 0.5*cm))

    doc.build(elements)
    buffer.seek(0)
    return buffer


def create_statement_pdf(title, subtitle, sections, footer_lines=None,
                         amount_headers=None):
    """تقرير قائمة مالية (ميزانية/دخل) بصيغة PDF عمودية.

    sections: قائمة من (عنوان القسم، rows=[(code, name, amount)], الإجمالي)
    footer_lines: قائمة أسطر ختامية اختيارية.
    amount_headers: عناوين أعمدة المبالغ (مثلاً عند المقارنة بين فترتين).
        عند تمريرها، يُتوقع أن يكون amount في كل صف قائمة من القيم.
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import io

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            leftMargin=1.6*cm, rightMargin=1.6*cm,
                            topMargin=1.8*cm, bottomMargin=1.8*cm)

    try:
        pdfmetrics.registerFont(TTFont('Arabic', 'C:/Windows/Fonts/arial.ttf'))
        font_name = 'Arabic'
    except Exception:
        font_name = 'Helvetica'

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle('ArTitle', parent=styles['Title'],
                              fontName=font_name, fontSize=17, spaceAfter=8,
                              alignment=1))
    styles.add(ParagraphStyle('ArSubtitle', parent=styles['Normal'],
                              fontName=font_name, fontSize=10, spaceAfter=8,
                              alignment=1, textColor=colors.grey))
    styles.add(ParagraphStyle('ArSection', parent=styles['Normal'],
                              fontName=font_name, fontSize=12, spaceAfter=6,
                              textColor=colors.HexColor('#4F46E5')))
    styles.add(ParagraphStyle('ArFooter', parent=styles['Normal'],
                              fontName=font_name, fontSize=10, spaceBefore=10,
                              alignment=1))

    elements = []
    elements += company_header_elements(font_name)
    elements.append(Paragraph(title, styles['ArTitle']))
    elements.append(Paragraph(subtitle, styles['ArSubtitle']))
    elements.append(Spacer(1, 0.4*cm))

    _cur_suffix = currency_suffix()
    if amount_headers:
        headers_row = ['#', 'البيان'] + [f'{h}{_cur_suffix}' for h in amount_headers]
        n_amounts = len(amount_headers)
        amount_w = max(2.6*cm, 8.8*cm / n_amounts)
        col_widths = [1.2*cm, 10.5*cm] + [amount_w] * n_amounts
    else:
        headers_row = ['#', 'البيان', f'المبلغ{_cur_suffix}']
        col_widths = [1.4*cm, 12*cm, 3.4*cm]

    for section_title, rows, total in sections:
        elements.append(Paragraph(section_title, styles['ArSection']))
        data = [headers_row]
        for i, (code, name, amount) in enumerate(rows, 1):
            label = f'{code} - {name}' if code else name
            if amount_headers:
                amounts = amount if isinstance(amount, (list, tuple)) else [amount]
                amounts = [f'{float(a or 0):,.2f}' for a in amounts]
                data.append([str(i), label] + amounts)
            else:
                data.append([str(i), label, f'{float(amount or 0):,.2f}'])
        if amount_headers:
            totals = total if isinstance(total, (list, tuple)) else [total]
            data.append(['', 'الإجمالي'] + [f'{float(t or 0):,.2f}' for t in totals])
        else:
            data.append(['', 'الإجمالي', f'{float(total or 0):,.2f}'])
        table = Table(data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), font_name),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),
            ('ALIGN', (2, 1), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2),
             [colors.white, colors.HexColor('#F8FAFC')]),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#F1F5F9')),
            ('FONTNAME', (0, -1), (-1, -1), font_name),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 0.5*cm))

    if footer_lines:
        elements.append(Paragraph('<br/>'.join(footer_lines),
                                  styles['ArFooter']))

    doc.build(elements)
    buffer.seek(0)
    return buffer


def create_financial_overview_pdf(title, subtitle, reports):
    """تقرير مالي شامل يجمع عدة قوائم مالية في مستند PDF واحد.

    reports: قائمة من (عنوان القائمة، amount_headers، sections=[(قسم، rows، إجمالي)]، footer)
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import io

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            leftMargin=1.6*cm, rightMargin=1.6*cm,
                            topMargin=1.8*cm, bottomMargin=1.8*cm)
    try:
        pdfmetrics.registerFont(TTFont('Arabic', 'C:/Windows/Fonts/arial.ttf'))
        font_name = 'Arabic'
    except Exception:
        font_name = 'Helvetica'

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle('ArTitle', parent=styles['Title'],
                              fontName=font_name, fontSize=17, spaceAfter=8,
                              alignment=1))
    styles.add(ParagraphStyle('ArSubtitle', parent=styles['Normal'],
                              fontName=font_name, fontSize=10, spaceAfter=8,
                              alignment=1, textColor=colors.grey))
    styles.add(ParagraphStyle('ArSection', parent=styles['Normal'],
                              fontName=font_name, fontSize=12, spaceAfter=6,
                              textColor=colors.HexColor('#4F46E5')))
    styles.add(ParagraphStyle('ArFooter', parent=styles['Normal'],
                              fontName=font_name, fontSize=10, spaceBefore=10,
                              alignment=1))
    styles.add(ParagraphStyle('ArStatement', parent=styles['Normal'],
                              fontName=font_name, fontSize=14, spaceBefore=12,
                              spaceAfter=4, textColor=colors.HexColor('#0F766E'),
                              alignment=1))

    elements = []
    elements += company_header_elements(font_name)
    elements.append(Paragraph(title, styles['ArTitle']))
    elements.append(Paragraph(subtitle, styles['ArSubtitle']))
    elements.append(Spacer(1, 0.4*cm))

    _cur_suffix = currency_suffix()
    for idx, (stmt_title, amount_headers, sections, footer_lines) in enumerate(reports):
        if idx:
            elements.append(Spacer(1, 0.5*cm))
        elements.append(Paragraph(stmt_title, styles['ArStatement']))
        if amount_headers:
            headers_row = ['#', 'البيان'] + [f'{h}{_cur_suffix}' for h in amount_headers]
            n_amounts = len(amount_headers)
            amount_w = max(2.6*cm, 8.8*cm / n_amounts)
            col_widths = [1.2*cm, 10.5*cm] + [amount_w] * n_amounts
        else:
            headers_row = ['#', 'البيان', f'المبلغ{_cur_suffix}']
            col_widths = [1.4*cm, 12*cm, 3.4*cm]
        for section_title, rows, total in sections:
            elements.append(Paragraph(section_title, styles['ArSection']))
            data = [headers_row]
            for i, (code, name, amount) in enumerate(rows, 1):
                label = f'{code} - {name}' if code else name
                if amount_headers:
                    amounts = amount if isinstance(amount, (list, tuple)) else [amount]
                    amounts = [f'{float(a or 0):,.2f}' for a in amounts]
                    data.append([str(i), label] + amounts)
                else:
                    data.append([str(i), label, f'{float(amount or 0):,.2f}'])
            if amount_headers:
                totals = total if isinstance(total, (list, tuple)) else [total]
                data.append(['', 'الإجمالي'] + [f'{float(t or 0):,.2f}' for t in totals])
            else:
                data.append(['', 'الإجمالي', f'{float(total or 0):,.2f}'])
            table = Table(data, colWidths=col_widths, repeatRows=1)
            table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), font_name),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('ALIGN', (0, 0), (0, -1), 'CENTER'),
                ('ALIGN', (2, 1), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -2),
                 [colors.white, colors.HexColor('#F8FAFC')]),
                ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#F1F5F9')),
                ('FONTNAME', (0, -1), (-1, -1), font_name),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ]))
            elements.append(table)
            elements.append(Spacer(1, 0.4*cm))
        if footer_lines:
            elements.append(Paragraph('<br/>'.join(footer_lines),
                                      styles['ArFooter']))

    doc.build(elements)
    buffer.seek(0)
    return buffer


def create_client_statement_pdf(data, date_from="", date_to=""):
    """PDF لكشف حساب عميل: فواتير (مدين) + مدفوعات (دائن) برصيد جاري."""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import io

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            leftMargin=1.2*cm, rightMargin=1.2*cm,
                            topMargin=1.8*cm, bottomMargin=1.8*cm)

    try:
        pdfmetrics.registerFont(TTFont('Arabic', 'C:/Windows/Fonts/arial.ttf'))
        font_name = 'Arabic'
    except Exception:
        font_name = 'Helvetica'

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle('ArTitle', parent=styles['Title'],
                              fontName=font_name, fontSize=17, spaceAfter=8,
                              alignment=1))
    styles.add(ParagraphStyle('ArSubtitle', parent=styles['Normal'],
                              fontName=font_name, fontSize=10, spaceAfter=8,
                              alignment=1, textColor=colors.grey))

    client = data['client']
    elements = []
    elements += company_header_elements(font_name)
    elements.append(Paragraph(f'كشف حساب عميل: {client.name}',
                              styles['ArTitle']))
    subtitle = f"تاريخ التقرير: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    if date_from or date_to:
        subtitle += f" — الفترة: {date_from or '...'} إلى {date_to or '...'}"
    elements.append(Paragraph(subtitle, styles['ArSubtitle']))
    elements.append(Spacer(1, 0.4*cm))

    headers = ['#', 'المرجع', 'التاريخ', 'البيان', f'مدين{currency_suffix()}',
               f'دائن{currency_suffix()}', f'الرصيد{currency_suffix()}']
    col_widths = [1*cm, 3.4*cm, 2.2*cm, 7.5*cm, 2.8*cm, 2.8*cm, 3*cm]
    table_style = TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), font_name),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('ALIGN', (4, 1), (6, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2),
         [colors.white, colors.HexColor('#F8FAFC')]),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#F1F5F9')),
        ('FONTNAME', (0, -1), (-1, -1), font_name),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ])

    rows_data = [headers]
    rows_data.append(['', 'الرصيد الافتتاحي', '', '', '', '',
                      f"{data['opening']:,.2f}"])
    for i, row in enumerate(data['rows'], 1):
        rows_data.append([
            str(i),
            row['ref'],
            row['date'].strftime('%Y-%m-%d') if row['date'] else '',
            row['notes'] or '',
            f"{row['debit']:,.2f}" if row['debit'] else '',
            f"{row['credit']:,.2f}" if row['credit'] else '',
            f"{row['running']:,.2f}",
        ])
    rows_data.append(['', 'الرصيد الختامي', '', '', '', '',
                      f"{data['closing']:,.2f}"])
    table = Table(rows_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(table_style)
    elements.append(table)

    summary = (f"إجمالي الفواتير: {data['total_debt']:,.2f} — "
               f"إجمالي المدفوعات: {data['total_paid']:,.2f} — "
               f"الرصيد المتبقي: {data['closing']:,.2f}")
    elements.append(Paragraph(summary, styles['ArSubtitle']))

    doc.build(elements)
    buffer.seek(0)
    return buffer


def create_aging_pdf(buckets, asof=None):
    """PDF لتقرير تقادم الديون: مصفوفة أعمار المستحقات حسب الفئات."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import io

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.8*cm, bottomMargin=1.8*cm)

    try:
        pdfmetrics.registerFont(TTFont('Arabic', 'C:/Windows/Fonts/arial.ttf'))
        font_name = 'Arabic'
    except Exception:
        font_name = 'Helvetica'

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle('ArTitle', parent=styles['Title'],
                              fontName=font_name, fontSize=17, spaceAfter=8,
                              alignment=1))
    styles.add(ParagraphStyle('ArSubtitle', parent=styles['Normal'],
                              fontName=font_name, fontSize=10, spaceAfter=8,
                              alignment=1, textColor=colors.grey))
    styles.add(ParagraphStyle('ArSection', parent=styles['Normal'],
                              fontName=font_name, fontSize=12, spaceAfter=6,
                              textColor=colors.HexColor('#4F46E5')))
    styles.add(ParagraphStyle('ArFooter', parent=styles['Normal'],
                              fontName=font_name, fontSize=10, spaceBefore=10,
                              alignment=1))

    _cur_suffix = currency_suffix()
    elements = []
    elements += company_header_elements(font_name)
    elements.append(Paragraph('تقرير تقادم الديون', styles['ArTitle']))
    asof_label = asof.strftime('%Y-%m-%d') if asof else ''
    elements.append(Paragraph(f'التاريخ المرجعي: {asof_label}',
                              styles['ArSubtitle']))
    elements.append(Spacer(1, 0.4*cm))

    headers = ['#', 'العميل', 'الهاتف', 'الأيام', f'إجمالي المديونية{_cur_suffix}',
               f'المدفوع{_cur_suffix}', f'المتبقي{_cur_suffix}']
    col_widths = [1*cm, 4.2*cm, 3.2*cm, 1.6*cm, 3.2*cm, 3.2*cm, 3.2*cm]

    grand_total = 0.0
    grand_clients = 0
    for key in ('current', '30', '60', '90'):
        bucket = buckets[key]
        if not bucket['clients']:
            continue
        elements.append(Paragraph(f"{bucket['label']} — {len(bucket['clients'])} عميل",
                                  styles['ArSection']))
        data = [headers]
        for i, (c, age, balance) in enumerate(bucket['clients'], 1):
            data.append([
                str(i), c.name, c.phone or '—', str(age),
                f"{float(c.total_debt or 0):,.2f}",
                f"{float(c.total_paid or 0):,.2f}",
                f"{balance:,.2f}",
            ])
        data.append(['', 'إجمالي الفئة', '', '',
                     f"{bucket['total']:,.2f}", '', f"{bucket['total']:,.2f}"])
        table = Table(data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), font_name),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),
            ('ALIGN', (3, 1), (6, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2),
             [colors.white, colors.HexColor('#F8FAFC')]),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#F1F5F9')),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 0.4*cm))
        grand_total += float(bucket['total'] or 0)
        grand_clients += len(bucket['clients'])

    if grand_clients:
        elements.append(Paragraph(
            f"إجمالي المستحقات المتأخرة: {grand_total:,.2f} {get_currency_short()}"
            f" — عدد العملاء: {grand_clients}",
            styles['ArFooter']))
    else:
        elements.append(Paragraph('لا توجد ديون متأخرة', styles['ArFooter']))

    doc.build(elements)
    buffer.seek(0)
    return buffer


def backup_database(app):
    import sqlite3 as _sqlite3
    db_path = os.path.join(app.instance_path, 'debtors.db')
    if not os.path.exists(db_path):
        return None

    backup_dir = os.path.join(os.path.dirname(app.instance_path), 'backups')
    os.makedirs(backup_dir, exist_ok=True)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    dst = os.path.join(backup_dir, f'db_backup_{timestamp}.db')

    try:
        src = _sqlite3.connect(db_path, timeout=5)
        src.execute('PRAGMA busy_timeout=5000')
        src.execute('PRAGMA wal_checkpoint(TRUNCATE)')
        dst_conn = _sqlite3.connect(dst, timeout=5)
        src.backup(dst_conn, pages=1000, name='main')
        dst_conn.close()
        src.close()
        app.logger.info(f"Database backed up to {dst}")
    except Exception:
        return None

    cutoff = datetime.now() - timedelta(days=30)
    for f in os.listdir(backup_dir):
        fpath = os.path.join(backup_dir, f)
        if os.path.isfile(fpath) and os.path.getmtime(fpath) < cutoff.timestamp():
            os.remove(fpath)
            app.logger.info(f"Old backup removed: {f}")

    export_dir = os.path.join(os.path.dirname(app.instance_path), 'exports')
    if os.path.exists(export_dir):
        export_cutoff = datetime.now() - timedelta(days=7)
        for f in os.listdir(export_dir):
            fpath = os.path.join(export_dir, f)
            if os.path.isfile(fpath) and os.path.getmtime(fpath) < export_cutoff.timestamp():
                os.remove(fpath)

    return dst


def cleanup_old_uploads(app):
    from app.models import Settings
    upload_dir = app.config.get('UPLOAD_FOLDER')
    if not upload_dir or not os.path.exists(upload_dir):
        return 0
    retention_days = int(Settings.get('upload_retention_days', '7'))
    cutoff = datetime.now() - timedelta(days=retention_days)
    removed = 0
    for f in os.listdir(upload_dir):
        fpath = os.path.join(upload_dir, f)
        if os.path.isfile(fpath) and os.path.getmtime(fpath) < cutoff.timestamp():
            try:
                os.remove(fpath)
                removed += 1
            except OSError:
                pass
    if removed:
        app.logger.info(f"Cleaned up {removed} old uploads (>{retention_days} days)")
    return removed


def export_excel(clients):
    wb = Workbook()
    ws_xl = wb.active
    ws_xl.title = 'تقرير المديونيات'
    ws_xl.sheet_view.rightToLeft = True

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='4F46E5')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='DDDDDD')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ['#', 'اسم العميل', 'رقم الهاتف', 'المديونية', 'المدفوع', 'المتبقي', 'الحالة']
    col_widths = [5, 25, 18, 15, 15, 15, 12]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws_xl.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
        ws_xl.column_dimensions[get_column_letter(col)].width = w

    ws_xl.row_dimensions[1].height = 30

    row_num = 2
    for c in clients:
        status_text = 'مدفوع' if c.status == 'paid' else 'مستحق'
        row_data = [row_num - 1, c.name, c.phone or '', c.total_debt, c.total_paid,
                    c.balance, status_text]
        for col, val in enumerate(row_data, 1):
            cell = ws_xl.cell(row=row_num, column=col, value=val)
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if col in (4, 5, 6):
                cell.number_format = '#,##0.00'
            if c.status == 'paid':
                cell.fill = PatternFill('solid', fgColor='F0FDF4')
        row_num += 1

    total_debt = sum(c.total_debt for c in clients)
    total_paid = sum(c.total_paid for c in clients)
    total_balance = sum(c.balance for c in clients)
    totals = ['', 'الإجمالي', '', total_debt, total_paid, total_balance, '']
    for col, val in enumerate(totals, 1):
        cell = ws_xl.cell(row=row_num, column=col, value=val)
        cell.font = Font(bold=True)
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        if col in (4, 5, 6):
            cell.number_format = '#,##0.00'
        cell.fill = PatternFill('solid', fgColor='F1F5F9')

    return wb


FIELD_MAP = {
    'name': {'اسم', 'name', 'عميل', 'client', 'اسم العميل', 'اسم العميل '},
    'phone': {'هاتف', 'phone', 'موبايل', 'تليفون', 'جوال', 'رقم الهاتف', 'رقم'},
    'total_debt': {'مديونية', 'debt', 'مبلغ', 'رصيد', 'المديونية', 'المبلغ', 'الرصيد', 'total_debt', 'amount'},
    'total_paid': {'مدفوع', 'paid', 'المدفوع', 'سداد', 'total_paid', 'المدفوعات'},
    'notes': {'ملاحظ', 'notes', 'بيان', 'ملاحظات', 'البيان'},
}


def parse_uploaded_file(file_storage):
    filename = (file_storage.filename or '').lower()
    if filename.endswith('.csv'):
        import csv, io as _io
        content = file_storage.read()
        file_storage.seek(0)
        for enc in ('utf-8-sig', 'utf-8', 'cp1256', 'cp1252', 'latin-1'):
            try:
                text = content.decode(enc)
                break
            except (UnicodeDecodeError, LookupError):
                continue
        else:
            text = content.decode('utf-8', errors='replace')
        reader = csv.reader(_io.StringIO(text))
        all_rows = [row for row in reader]
        if not all_rows:
            return None
        headers = [str(h).strip() for h in all_rows[0]]
        data = all_rows[1:]
    else:
        import openpyxl
        try:
            wb = openpyxl.load_workbook(file_storage, read_only=True, data_only=True)
            ws = wb.active
            all_rows = list(ws.iter_rows(values_only=True))
            wb.close()
        except Exception:
            try:
                wb.close()
            except Exception:
                pass
            return None
        if not all_rows:
            return None
        headers = [str(h).strip() if h else '' for h in all_rows[0]]
        data = [list(r) for r in all_rows[1:]]

    return {'headers': headers, 'data': data, 'row_count': len(data)}


def auto_detect_columns(headers):
    mapping = {}
    for field, keywords in FIELD_MAP.items():
        for idx, h in enumerate(headers):
            h_lower = h.lower().strip()
            if any(k in h_lower for k in keywords):
                mapping[field] = idx
                break
    return mapping


def create_sample_template():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = 'بيانات العملاء'
    ws.sheet_view.rightToLeft = True

    thin = Side(style='thin', color='DDDDDD')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', fgColor='4F46E5')

    headers = ['اسم العميل', 'رقم الهاتف', 'المديونية', 'المدفوع', 'ملاحظات']
    col_widths = [25, 18, 15, 15, 20]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 30

    samples = [
        ['أحمد محمد', '01012345678', 5000, 2000, 'عميل قديم'],
        ['سارة علي', '01098765432', 3500, 0, ''],
        ['خالد حسن', '01155566677', 12000, 8000, 'vip'],
    ]
    for r, row_data in enumerate(samples, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')

    return wb


def validate_import_rows(data, mapping):
    validated = []
    for row_num, row in enumerate(data, 1):
        errors = []
        name_idx = mapping.get('name')
        phone_idx = mapping.get('phone')
        debt_idx = mapping.get('total_debt')
        paid_idx = mapping.get('total_paid')
        notes_idx = mapping.get('notes')

        name = ''
        if name_idx is not None and name_idx < len(row) and row[name_idx]:
            name = str(row[name_idx]).strip()
        if not name:
            errors.append('اسم العميل مطلوب')

        phone = ''
        if phone_idx is not None and phone_idx < len(row) and row[phone_idx]:
            phone = str(row[phone_idx]).strip().replace(' ', '').replace('-', '')

        debt_val = 0.0
        if debt_idx is not None and debt_idx < len(row) and row[debt_idx]:
            try:
                debt_val = float(str(row[debt_idx]).replace(',', '').strip())
                if debt_val < 0:
                    errors.append('المديونية لا يمكن أن تكون سالبة')
            except (ValueError, TypeError):
                errors.append(f'قيمة المديونية غير صحيحة: {row[debt_idx]}')

        paid_val = 0.0
        if paid_idx is not None and paid_idx < len(row) and row[paid_idx]:
            try:
                paid_val = float(str(row[paid_idx]).replace(',', '').strip())
                if paid_val < 0:
                    errors.append('المدفوع لا يمكن أن يكون سالبا')
            except (ValueError, TypeError):
                errors.append(f'قيمة المدفوع غير صحيحة: {row[paid_idx]}')

        notes_val = ''
        if notes_idx is not None and notes_idx < len(row) and row[notes_idx]:
            notes_val = str(row[notes_idx]).strip()

        validated.append({
            'row_num': row_num,
            'name': name,
            'phone': phone,
            'total_debt': debt_val,
            'total_paid': paid_val,
            'notes': notes_val,
            'errors': errors,
            'valid': len(errors) == 0 and bool(name),
        })

    return validated
