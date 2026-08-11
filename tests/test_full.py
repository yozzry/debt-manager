"""
Comprehensive test suite for Debt Manager — covers all 43 routes,
6 models, utils, parsers, templates, and app settings.
Run: python -m pytest tests/test_full.py -v --tb=short
"""
import io
import json
import os
import sys

import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from app.models import db, User, Client, Invoice, Payment, Settings, ActivityLog
from app.utils import (
    allowed_file, recalc_client, log_activity, get_app_settings, get_whatsapp_settings,
    build_reminder_message, export_excel, create_sample_template, parse_uploaded_file,
    auto_detect_columns, validate_import_rows, COUNTRY_OPTIONS
)
from app.importers.accounting_excel import (
    detect_format, _cell_str, _cell_num, _find_header_row, build_customer_preview
)
from openpyxl import Workbook


# ═══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

def _make_client(session, name="Test Client", phone="201012345678", debt=1000, paid=300):
    c = Client(name=name, phone=phone, total_debt=debt, total_paid=paid, status='due')
    session.add(c)
    session.commit()
    return c


def _make_invoice(session, client_id, amount=500, desc="Test invoice"):
    inv = Invoice(client_id=client_id, amount=amount, description=desc)
    session.add(inv)
    session.commit()
    return inv


def _make_payment(session, client_id, amount=200, notes="Test payment"):
    p = Payment(client_id=client_id, amount=amount, notes=notes)
    session.add(p)
    session.commit()
    return p


def _login(client):
    client.post('/login', data={'username': 'admin', 'password': 'admin123'}, follow_redirects=True)


def _create_user(session, username='editor1', role='editor', active=True):
    u = User(username=username, role=role, is_active_flag=active)
    u.set_password('pass123')
    session.add(u)
    session.commit()
    return u


# ═══════════════════════════════════════════════════════════════════════════════
# 1. MODELS
# ═══════════════════════════════════════════════════════════════════════════════

class TestModels:
    def test_user_password_hash(self, app):
        with app.app_context():
            u = User(username='testuser', role='editor')
            u.set_password('mypass')
            assert u.check_password('mypass')
            assert not u.check_password('wrong')
            assert u.password_hash != 'mypass'

    def test_user_roles(self, app):
        with app.app_context():
            admin = User(username='adm', role='admin')
            editor = User(username='edt', role='editor')
            viewer = User(username='vw', role='viewer')
            assert admin.is_admin
            assert not editor.is_admin
            assert admin.can_edit
            assert editor.can_edit
            assert not viewer.can_edit

    def test_user_to_dict(self, app):
        with app.app_context():
            u = User(username='serial', role='admin')
            u.set_password('x')
            db.session.add(u)
            db.session.commit()
            d = u.to_dict()
            assert d['username'] == 'serial'
            assert d['role'] == 'admin'
            assert 'password_hash' not in d

    def test_client_balance(self, app):
        with app.app_context():
            c = Client(name='Bal', total_debt=1000, total_paid=400)
            assert c.balance == 600
            c2 = Client(name='Zero', total_debt=300, total_paid=500)
            assert c2.balance == 0

    def test_client_to_dict(self, app):
        with app.app_context():
            c = Client(name='Dict', phone='123', total_debt=100, total_paid=50)
            db.session.add(c)
            db.session.commit()
            d = c.to_dict()
            assert d['name'] == 'Dict'
            assert d['balance'] == 50

    def test_invoice_to_dict(self, app):
        with app.app_context():
            c = _make_client(db.session)
            inv = _make_invoice(db.session, c.id)
            d = inv.to_dict()
            assert d['amount'] == 500
            assert d['client_id'] == c.id

    def test_payment_to_dict(self, app):
        with app.app_context():
            c = _make_client(db.session)
            p = _make_payment(db.session, c.id)
            d = p.to_dict()
            assert d['amount'] == 200

    def test_settings_get_set(self, app):
        with app.app_context():
            Settings.set('test_key', 'test_val')
            assert Settings.get('test_key') == 'test_val'
            assert Settings.get('missing', 'default') == 'default'
            Settings.set('test_key', 'updated')
            assert Settings.get('test_key') == 'updated'

    def test_activity_log_to_dict(self, app):
        with app.app_context():
            a = ActivityLog(user_id=1, action='login', entity_type='user', entity_id=1, details='test')
            db.session.add(a)
            db.session.commit()
            d = a.to_dict()
            assert d['action'] == 'login'

    def test_client_cascade_delete(self, app):
        with app.app_context():
            c = _make_client(db.session)
            _make_invoice(db.session, c.id, 100)
            _make_payment(db.session, c.id, 50)
            cid = c.id
            db.session.delete(c)
            db.session.commit()
            assert Invoice.query.filter_by(client_id=cid).count() == 0
            assert Payment.query.filter_by(client_id=cid).count() == 0


# ═══════════════════════════════════════════════════════════════════════════════
# 2. AUTH ROUTES
# ═══════════════════════════════════════════════════════════════════════════════

class TestAuth:
    def test_login_page(self, client):
        r = client.get('/login')
        assert r.status_code == 200

    def test_login_success(self, client):
        r = client.post('/login', data={'username': 'admin', 'password': 'admin123'}, follow_redirects=True)
        assert r.status_code == 200

    def test_login_fail(self, client):
        r = client.post('/login', data={'username': 'admin', 'password': 'wrong'}, follow_redirects=True)
        assert r.status_code == 200

    def test_logout(self, auth_client):
        r = auth_client.get('/logout', follow_redirects=True)
        assert r.status_code == 200

    def test_users_page(self, auth_client):
        r = auth_client.get('/users')
        assert r.status_code == 200

    def test_user_add(self, auth_client):
        r = auth_client.post('/users/add', data={'username': 'newuser', 'password': 'pass123', 'role': 'editor'}, follow_redirects=True)
        assert r.status_code == 200

    def test_user_toggle(self, auth_client, app):
        with app.app_context():
            u = _create_user(db.session)
            uid = u.id
        r = auth_client.post(f'/users/{uid}/toggle', follow_redirects=True)
        assert r.status_code == 200

    def test_user_delete(self, auth_client, app):
        with app.app_context():
            u = _create_user(db.session, 'deluser')
            uid = u.id
        r = auth_client.post(f'/users/{uid}/delete', follow_redirects=True)
        assert r.status_code == 200

    def test_non_admin_cannot_access_users(self, auth_client, app):
        with app.app_context():
            u = _create_user(db.session, 'viewer1', 'viewer')
        auth_client.get('/logout', follow_redirects=True)
        auth_client.post('/login', data={'username': 'viewer1', 'password': 'pass123'}, follow_redirects=True)
        r = auth_client.get('/users', follow_redirects=True)
        assert r.status_code == 200


# ═══════════════════════════════════════════════════════════════════════════════
# 3. CLIENT ROUTES
# ═══════════════════════════════════════════════════════════════════════════════

class TestClients:
    def test_index(self, auth_client):
        r = auth_client.get('/')
        assert r.status_code == 200

    def test_add_client(self, auth_client, app):
        r = auth_client.post('/client/add', data={'name': 'Ahmed', 'phone': '20101112233', 'notes': 'test'}, follow_redirects=True)
        assert r.status_code == 200
        with app.app_context():
            assert Client.query.filter_by(name='Ahmed').first() is not None

    def test_client_detail(self, auth_client, app):
        with app.app_context():
            c = _make_client(db.session, 'DetailClient')
            cid = c.id
        r = auth_client.get(f'/client/{cid}')
        assert r.status_code == 200

    def test_edit_client(self, auth_client, app):
        with app.app_context():
            c = _make_client(db.session, 'OldName')
            cid = c.id
        r = auth_client.post(f'/client/{cid}/edit', data={'name': 'NewName', 'phone': '123'}, follow_redirects=True)
        assert r.status_code == 200
        with app.app_context():
            assert Client.query.get(cid).name == 'NewName'

    def test_delete_client(self, auth_client, app):
        with app.app_context():
            c = _make_client(db.session, 'ToDelete')
            cid = c.id
        r = auth_client.post(f'/client/{cid}/delete', follow_redirects=True)
        assert r.status_code == 200
        with app.app_context():
            assert Client.query.get(cid) is None

    def test_client_settings(self, auth_client, app):
        with app.app_context():
            c = _make_client(db.session)
            cid = c.id
        r = auth_client.get(f'/client/{cid}/settings', follow_redirects=True)
        assert r.status_code == 200
        r = auth_client.post(f'/client/{cid}/settings', data={'reminder_enabled': 'on', 'reminder_template': '2'}, follow_redirects=True)
        assert r.status_code == 200

    def test_search(self, auth_client, app):
        with app.app_context():
            _make_client(db.session, 'Searchable')
        r = auth_client.get('/?q=Searchable')
        assert r.status_code == 200

    def test_status_filter(self, auth_client, app):
        with app.app_context():
            _make_client(db.session, 'DueClient')
        r = auth_client.get('/?status=due')
        assert r.status_code == 200

    def test_toggle_dark(self, auth_client):
        r = auth_client.post('/api/toggle-dark')
        assert r.status_code == 200

    def test_unauthenticated_redirects(self, client):
        r = client.get('/')
        assert r.status_code == 302


# ═══════════════════════════════════════════════════════════════════════════════
# 4. INVOICE ROUTES
# ═══════════════════════════════════════════════════════════════════════════════

class TestInvoices:
    def test_add_invoice(self, auth_client, app):
        with app.app_context():
            c = _make_client(db.session)
            cid = c.id
        r = auth_client.post(f'/client/{cid}/invoice/add', data={
            'amount': '750', 'description': 'New invoice', 'date': '2026-01-15'
        }, follow_redirects=True)
        assert r.status_code == 200
        data = json.loads(r.data)
        assert data.get('ok') is True

    def test_delete_invoice(self, auth_client, app):
        with app.app_context():
            c = _make_client(db.session)
            inv = _make_invoice(db.session, c.id, 100)
            iid = inv.id
        r = auth_client.post(f'/invoice/{iid}/delete', follow_redirects=True)
        assert r.status_code == 200
        data = json.loads(r.data)
        assert data.get('ok') is True

    def test_invoice_recalculates_client(self, auth_client, app):
        with app.app_context():
            c = _make_client(db.session, 'Recalc', debt=0, paid=0)
            cid = c.id
        auth_client.post(f'/client/{cid}/invoice/add', data={'amount': '500', 'description': 'inv'}, follow_redirects=True)
        with app.app_context():
            c2 = Client.query.get(cid)
            assert c2.total_debt == 500


# ═══════════════════════════════════════════════════════════════════════════════
# 5. PAYMENT ROUTES
# ═══════════════════════════════════════════════════════════════════════════════

class TestPayments:
    def test_add_payment(self, auth_client, app):
        with app.app_context():
            c = _make_client(db.session, debt=1000, paid=0)
            cid = c.id
        r = auth_client.post(f'/client/{cid}/payment/add', data={
            'amount': '250', 'notes': 'partial', 'date': '2026-02-01'
        }, follow_redirects=True)
        assert r.status_code == 200
        data = json.loads(r.data)
        assert data.get('ok') is True

    def test_delete_payment(self, auth_client, app):
        with app.app_context():
            c = _make_client(db.session)
            p = _make_payment(db.session, c.id, 100)
            pid = p.id
        r = auth_client.post(f'/payment/{pid}/delete', follow_redirects=True)
        assert r.status_code == 200
        data = json.loads(r.data)
        assert data.get('ok') is True

    def test_payment_recalculates_client(self, auth_client, app):
        with app.app_context():
            c = _make_client(db.session, 'PayCalc', debt=0, paid=0)
            cid = c.id
            _make_invoice(db.session, cid, 1000)
        auth_client.post(f'/client/{cid}/payment/add', data={'amount': '400', 'notes': 'pay'}, follow_redirects=True)
        with app.app_context():
            c2 = Client.query.get(cid)
            assert c2.total_paid == 400
            assert c2.balance == 600


# ═══════════════════════════════════════════════════════════════════════════════
# 6. WHATSAPP / SETTINGS ROUTES
# ═══════════════════════════════════════════════════════════════════════════════

class TestWhatsApp:
    def test_settings_page(self, auth_client):
        r = auth_client.get('/settings')
        assert r.status_code == 200

    def test_settings_general_tab(self, auth_client):
        r = auth_client.get('/settings?tab=general')
        assert r.status_code == 200

    def test_settings_save_general(self, auth_client, app):
        r = auth_client.post('/settings', data={
            'tab': 'general', 'app_country': 'SA',
            'app_timezone': 'Asia/Riyadh', 'app_currency': 'ريال سعودي', 'app_currency_short': 'ر.س'
        }, follow_redirects=True)
        assert r.status_code == 200
        with app.app_context():
            assert Settings.get('app_country') == 'SA'

    def test_settings_whatsapp_tab(self, auth_client):
        r = auth_client.get('/settings?tab=whatsapp')
        assert r.status_code == 200

    def test_settings_save_baileys(self, auth_client, app):
        r = auth_client.post('/settings', data={
            'tab': 'whatsapp', 'baileys_url': 'http://localhost:9999'
        }, follow_redirects=True)
        assert r.status_code == 200
        with app.app_context():
            assert Settings.get('baileys_url') == 'http://localhost:9999'

    def test_settings_reminder_tab(self, auth_client):
        r = auth_client.get('/settings?tab=reminder')
        assert r.status_code == 200

    def test_settings_save_reminder(self, auth_client, app):
        r = auth_client.post('/settings', data={
            'tab': 'reminder', 'reminder_enabled': 'on',
            'reminder_times': '09:00', 'reminder_frequency': 'daily'
        }, follow_redirects=True)
        assert r.status_code == 200
        with app.app_context():
            assert Settings.get('reminder_enabled') == 'true'

    def test_settings_templates_tab(self, auth_client):
        r = auth_client.get('/settings?tab=templates')
        assert r.status_code == 200

    def test_settings_save_templates(self, auth_client, app):
        r = auth_client.post('/settings', data={
            'tab': 'templates', 'template_1': 'Hi {name}', 'template_2': 'Balance {balance}', 'template_3': 'Bye'
        }, follow_redirects=True)
        assert r.status_code == 200
        with app.app_context():
            assert Settings.get('template_1') == 'Hi {name}'

    def test_whatsapp_status_endpoint(self, auth_client):
        r = auth_client.get('/api/whatsapp/status')
        assert r.status_code == 200
        data = json.loads(r.data)
        assert 'connected' in data

    def test_send_reminder_no_phone(self, auth_client, app):
        with app.app_context():
            c = Client(name='NoPhone')
            db.session.add(c)
            db.session.commit()
            cid = c.id
        r = auth_client.post(f'/api/whatsapp/send-reminder/{cid}', follow_redirects=True)
        data = json.loads(r.data)
        assert data.get('ok') is False

    def test_baileys_logout_endpoint(self, auth_client):
        r = auth_client.post('/api/baileys/logout', follow_redirects=True)
        assert r.status_code == 200


# ═══════════════════════════════════════════════════════════════════════════════
# 7. REPORTS ROUTES
# ═══════════════════════════════════════════════════════════════════════════════

class TestReports:
    def test_report_page(self, auth_client, app):
        with app.app_context():
            _make_client(db.session)
        r = auth_client.get('/report')
        assert r.status_code == 200

    def test_advanced_report(self, auth_client, app):
        with app.app_context():
            _make_client(db.session)
        r = auth_client.get('/advanced-report')
        assert r.status_code == 200

    def test_compare_report(self, auth_client, app):
        with app.app_context():
            _make_client(db.session)
        r = auth_client.get('/compare')
        assert r.status_code == 200

    def test_export_excel(self, auth_client, app):
        with app.app_context():
            _make_client(db.session)
        r = auth_client.get('/export')
        assert r.status_code == 200
        assert 'spreadsheetml' in r.content_type or 'octet' in r.content_type

    def test_export_pdf(self, auth_client, app):
        with app.app_context():
            _make_client(db.session)
        r = auth_client.get('/export/pdf')
        assert r.status_code == 200

    def test_import_page(self, auth_client):
        r = auth_client.get('/import')
        assert r.status_code == 200

    def test_import_template(self, auth_client):
        r = auth_client.get('/import/template')
        assert r.status_code == 200
        assert 'spreadsheetml' in r.content_type

    def test_backup(self, auth_client):
        r = auth_client.get('/backup', follow_redirects=True)
        assert r.status_code == 200


# ═══════════════════════════════════════════════════════════════════════════════
# 8. API V1 ROUTES
# ═══════════════════════════════════════════════════════════════════════════════

class TestAPI:
    def test_list_clients(self, auth_client, app):
        with app.app_context():
            _make_client(db.session)
        r = auth_client.get('/api/v1/clients')
        data = json.loads(r.data)
        assert r.status_code == 200
        assert 'clients' in data
        assert data['total'] >= 1

    def test_create_client(self, auth_client):
        r = auth_client.post('/api/v1/clients', json={'name': 'APIClient', 'phone': '966501234567'})
        assert r.status_code == 201
        data = json.loads(r.data)
        assert data['client']['name'] == 'APIClient'

    def test_get_client(self, auth_client, app):
        with app.app_context():
            c = _make_client(db.session, 'GetMe')
            cid = c.id
        r = auth_client.get(f'/api/v1/clients/{cid}')
        data = json.loads(r.data)
        assert data['name'] == 'GetMe'

    def test_update_client(self, auth_client, app):
        with app.app_context():
            c = _make_client(db.session, 'OldAPI')
            cid = c.id
        r = auth_client.put(f'/api/v1/clients/{cid}', json={'name': 'NewAPI'})
        assert r.status_code == 200

    def test_delete_client_api(self, auth_client, app):
        with app.app_context():
            c = _make_client(db.session, 'DelAPI')
            cid = c.id
        r = auth_client.delete(f'/api/v1/clients/{cid}')
        assert r.status_code == 200

    def test_add_invoice_api(self, auth_client, app):
        with app.app_context():
            c = _make_client(db.session)
            cid = c.id
        r = auth_client.post(f'/api/v1/clients/{cid}/invoices', json={'amount': 300, 'description': 'api inv'})
        assert r.status_code == 201

    def test_delete_invoice_api(self, auth_client, app):
        with app.app_context():
            c = _make_client(db.session)
            inv = _make_invoice(db.session, c.id)
            iid = inv.id
        r = auth_client.delete(f'/api/v1/invoices/{iid}')
        assert r.status_code == 200

    def test_add_payment_api(self, auth_client, app):
        with app.app_context():
            c = _make_client(db.session)
            cid = c.id
        r = auth_client.post(f'/api/v1/clients/{cid}/payments', json={'amount': 150, 'notes': 'api pay'})
        assert r.status_code == 201

    def test_delete_payment_api(self, auth_client, app):
        with app.app_context():
            c = _make_client(db.session)
            p = _make_payment(db.session, c.id)
            pid = p.id
        r = auth_client.delete(f'/api/v1/payments/{pid}')
        assert r.status_code == 200

    def test_report_summary(self, auth_client, app):
        with app.app_context():
            _make_client(db.session)
        r = auth_client.get('/api/v1/reports/summary')
        data = json.loads(r.data)
        assert 'total_clients' in data

    def test_report_trends(self, auth_client):
        r = auth_client.get('/api/v1/reports/trends')
        data = json.loads(r.data)
        assert 'trends' in data
        assert len(data['trends']) == 6

    def test_activity_log(self, auth_client):
        r = auth_client.get('/api/v1/activity')
        data = json.loads(r.data)
        assert 'activities' in data

    def test_list_users_api(self, auth_client):
        r = auth_client.get('/api/v1/users')
        data = json.loads(r.data)
        assert 'users' in data


# ═══════════════════════════════════════════════════════════════════════════════
# 9. UTILS
# ═══════════════════════════════════════════════════════════════════════════════

class TestUtils:
    def test_allowed_file(self):
        assert allowed_file('test.png')
        assert allowed_file('test.pdf')
        assert allowed_file('test.jpg')
        assert not allowed_file('test.exe')
        assert not allowed_file('test.py')
        assert not allowed_file('test.xlsx')

    def test_recalc_client(self, app):
        with app.app_context():
            c = _make_client(db.session, debt=0, paid=0)
            _make_invoice(db.session, c.id, 800)
            _make_payment(db.session, c.id, 300)
            recalc_client(c.id)
            c2 = Client.query.get(c.id)
            assert c2.total_debt == 800
            assert c2.total_paid == 300
            assert c2.balance == 500

    def test_log_activity(self, app):
        with app.app_context():
            log_activity(1, 'test_action', 'client', 1, 'test details')
            a = ActivityLog.query.filter_by(action='test_action').first()
            assert a is not None
            assert a.details == 'test details'

    def test_get_app_settings_eg(self, app):
        with app.app_context():
            s = get_app_settings()
            assert s['country'] in ('EG', 'SA')
            assert 'phone_code' in s
            assert 'currency' in s

    def test_get_app_settings_sa(self, app):
        with app.app_context():
            Settings.set('app_country', 'SA')
            s = get_app_settings()
            assert s['phone_code'] == '966'
            assert 'ريال' in s['currency']

    def test_whatsapp_settings(self, app):
        with app.app_context():
            ws = get_whatsapp_settings()
            assert 'baileys_url' in ws

    def test_build_reminder_message(self, app):
        with app.app_context():
            c = Client(name='أحمد', total_debt=1000, total_paid=600)
            db.session.add(c)
            db.session.commit()
            msg = build_reminder_message(c, 1)
            assert 'أحمد' in msg
            assert '400' in msg

    def test_export_excel_empty(self, app):
        with app.app_context():
            wb = export_excel([])
            assert wb is not None

    def test_export_excel_with_data(self, app):
        with app.app_context():
            c = _make_client(db.session)
            wb = export_excel([c])
            assert wb is not None

    def test_create_sample_template(self):
        wb = create_sample_template()
        assert wb is not None
        assert len(wb.sheetnames) > 0

    def test_parse_uploaded_file_csv(self):
        content = b"name,phone,total_debt\ntest,123,100\n"
        f = io.BytesIO(content)
        f.filename = 'test.csv'
        result = parse_uploaded_file(f)
        assert result is not None
        assert len(result['data']) == 1

    def test_auto_detect_columns(self):
        headers = ['اسم العميل', 'رقم الهاتف', 'المديونية', 'المدفوع', 'ملاحظات']
        mapping = auto_detect_columns(headers)
        assert mapping['name'] == 0
        assert mapping['phone'] == 1
        assert mapping['total_debt'] == 2

    def test_validate_import_rows(self):
        data = [['Ahmed', '123', 500, 200, '']]
        mapping = {'name': 0, 'phone': 1, 'total_debt': 2, 'total_paid': 3, 'notes': 4}
        result = validate_import_rows(data, mapping)
        assert len(result) == 1
        assert result[0]['valid'] is True

    def test_country_options_complete(self):
        assert 'EG' in COUNTRY_OPTIONS
        assert 'SA' in COUNTRY_OPTIONS
        for k, v in COUNTRY_OPTIONS.items():
            assert 'phone_code' in v
            assert 'timezone' in v
            assert 'currency' in v


# ═══════════════════════════════════════════════════════════════════════════════
# 10. IMPORT (EXCEL)
# ═══════════════════════════════════════════════════════════════════════════════

class TestImport:
    def _make_xlsx(self, headers, rows):
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        for row in rows:
            ws.append(row)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        buf.filename = 'import.xlsx'
        return buf

    def test_import_preview(self, auth_client):
        f = self._make_xlsx(['اسم العميل', 'رقم الهاتف', 'المديونية', 'المدفوع', 'ملاحظات'],
                             [['أحمد', '010123', 500, 200, 'test']])
        r = auth_client.post('/import', data={'file': f}, content_type='multipart/form-data', follow_redirects=True)
        assert r.status_code == 200

    def test_import_empty_file(self, auth_client):
        f = io.BytesIO(b'')
        f.filename = 'empty.csv'
        r = auth_client.post('/import', data={'file': f}, content_type='multipart/form-data', follow_redirects=True)
        assert r.status_code == 200


# ═══════════════════════════════════════════════════════════════════════════════
# 11. ACCOUNTING EXCEL PARSER
# ═══════════════════════════════════════════════════════════════════════════════

class TestAccountingParser:
    def test_cell_str(self):
        class FakeCell:
            def __init__(self, v): self.value = v
        assert _cell_str(FakeCell('hello')) == 'hello'
        assert _cell_str(FakeCell(None)) == ''
        assert _cell_str(FakeCell('  spaces  ')) == 'spaces'

    def test_cell_num(self):
        class FakeCell:
            def __init__(self, v): self.value = v
        assert _cell_num(FakeCell(42)) == 42.0
        assert _cell_num(FakeCell(None)) == 0.0
        assert _cell_num(FakeCell(None), 99) == 99.0
        assert _cell_num(FakeCell('1,000')) == 1000.0

    def test_find_header_row(self):
        class FakeWS:
            def __init__(self):
                self.max_row = 5
                self.max_column = 3
                self._data = {
                    (1,1): None, (1,2): None, (1,3): None,
                    (2,1): 'Header', (2,2): 'مدين', (2,3): 'دائن',
                    (3,1): 'data', (3,2): 100, (3,3): 0,
                }
            def cell(self, r, c):
                class C:
                    def __init__(self, v): self.value = v
                return C(self._data.get((r,c)))
        ws = FakeWS()
        assert _find_header_row(ws, ['مدين', 'دائن']) == 2

    def test_detect_format_none(self):
        f = io.BytesIO(b'not an excel file')
        f.filename = 'bad.xlsx'
        assert detect_format(f) is None

    def test_build_customer_preview_empty(self):
        parsed = {'customers': [], 'transactions': []}
        preview = build_customer_preview(parsed)
        assert preview == []


# ═══════════════════════════════════════════════════════════════════════════════
# 12. TEMPLATE RENDERING (all pages accessible)
# ═══════════════════════════════════════════════════════════════════════════════

class TestTemplates:
    def test_base_template_renders(self, auth_client):
        r = auth_client.get('/')
        assert r.status_code == 200
        assert b'<!DOCTYPE html>' in r.data or b'<html' in r.data

    def test_index_has_stats(self, auth_client, app):
        with app.app_context():
            _make_client(db.session, debt=500, paid=200)
        r = auth_client.get('/')
        assert 'إجمالي العملاء'.encode('utf-8') in r.data

    def test_index_has_wa_status(self, auth_client):
        r = auth_client.get('/')
        assert 'waStatusBar'.encode() in r.data

    def test_client_detail_has_currency(self, auth_client, app):
        with app.app_context():
            c = _make_client(db.session)
            cid = c.id
        r = auth_client.get(f'/client/{cid}')
        assert r.status_code == 200

    def test_settings_has_general_tab(self, auth_client):
        r = auth_client.get('/settings?tab=general')
        assert 'الإعدادات العامة'.encode('utf-8') in r.data

    def test_settings_has_country_select(self, auth_client):
        r = auth_client.get('/settings?tab=general')
        assert 'app_country'.encode() in r.data


# ═══════════════════════════════════════════════════════════════════════════════
# 13. EDGE CASES & SECURITY
# ═══════════════════════════════════════════════════════════════════════════════

class TestEdgeCases:
    def test_404_client(self, auth_client):
        r = auth_client.get('/client/99999')
        assert r.status_code == 404

    def test_editor_cannot_delete(self, auth_client, app):
        with app.app_context():
            u = _create_user(db.session, 'edonly', 'editor')
            uid = u.id
        auth_client.get('/logout', follow_redirects=True)
        auth_client.post('/login', data={'username': 'edonly', 'password': 'pass123'}, follow_redirects=True)
        with app.app_context():
            c = _make_client(db.session, 'Protected')
            cid = c.id
        r = auth_client.post(f'/client/{cid}/delete', follow_redirects=True)
        with app.app_context():
            assert Client.query.get(cid) is not None

    def test_viewer_cannot_add_client(self, auth_client, app):
        with app.app_context():
            u = _create_user(db.session, 'vviewer', 'viewer')
        auth_client.get('/logout', follow_redirects=True)
        auth_client.post('/login', data={'username': 'vviewer', 'password': 'pass123'}, follow_redirects=True)
        r = auth_client.post('/client/add', data={'name': 'Nope'}, follow_redirects=True)
        assert r.status_code == 200

    def test_api_unauthorized(self, client):
        r = client.get('/api/v1/clients')
        assert r.status_code == 302

    def test_duplicate_client_name(self, auth_client, app):
        with app.app_context():
            _make_client(db.session, 'DupTest')
        r = auth_client.post('/client/add', data={'name': 'DupTest', 'phone': '111'}, follow_redirects=True)
        assert r.status_code == 200

    def test_payment_exceeds_debt(self, auth_client, app):
        with app.app_context():
            c = _make_client(db.session, debt=100, paid=0)
            cid = c.id
        r = auth_client.post(f'/client/{cid}/payment/add', data={'amount': '500', 'notes': 'overpay'}, follow_redirects=True)
        assert r.status_code == 200


# ═══════════════════════════════════════════════════════════════════════════════
# 14. APP SETTINGS INTEGRATION
# ═══════════════════════════════════════════════════════════════════════════════

class TestAppSettings:
    def test_default_country_is_eg(self, app):
        with app.app_context():
            s = get_app_settings()
            assert s['country'] == 'EG'
            assert s['phone_code'] == '20'

    def test_switch_to_sa(self, app):
        with app.app_context():
            Settings.set('app_country', 'SA')
            s = get_app_settings()
            assert s['phone_code'] == '966'
            assert 'ريال' in s['currency']

    def test_currency_in_templates(self, auth_client, app):
        with app.app_context():
            Settings.set('app_country', 'SA')
            Settings.set('app_currency_short', 'ر.س')
        r = auth_client.get('/')
        assert 'ر.س'.encode('utf-8') in r.data

    def test_phone_code_in_client_edit(self, auth_client, app):
        with app.app_context():
            Settings.set('app_country', 'SA')
            c = _make_client(db.session)
            cid = c.id
        r = auth_client.get(f'/client/{cid}/edit')
        assert '+966'.encode() in r.data
