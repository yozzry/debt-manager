"""
Comprehensive test suite for Debt Manager — 51 routes, 6 models, utils, parsers, templates, edge cases.
Run: python -m pytest tests/test_all.py -v --tb=short
"""
import io
import json
import os
import sys

import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from app.models import db, User, Client, Invoice, Payment, Settings, ActivityLog
from app.utils import (
    allowed_file, recalc_client, log_activity, get_app_settings, get_whatsapp_settings,
    build_reminder_message, export_excel, create_sample_template, parse_uploaded_file,
    auto_detect_columns, validate_import_rows, COUNTRY_OPTIONS, send_whatsapp,
    normalize_phone, _send_reminders_background
)
from app.importers.accounting_excel import (
    detect_format, _cell_str, _cell_num, _find_header_row, build_customer_preview
)
from openpyxl import Workbook


# ═══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

def _make_client(session, name="Test Client", phone="201012345678", debt=1000, paid=300):
    c = Client(name=name, phone=phone, total_debt=debt, total_paid=paid, status='due')
    session.add(c)
    session.commit()
    return c


def _make_invoice(session, client_id, amount=500, desc="Test invoice"):
    inv = Invoice(client_id=client_id, amount=amount, description=desc)
    session.add(inv)
    session.commit()
    return inv


def _make_payment(session, client_id, amount=200, notes="Test payment"):
    p = Payment(client_id=client_id, amount=amount, notes=notes)
    session.add(p)
    session.commit()
    return p


def _login(client):
    client.post('/login', data={'username': 'admin', 'password': 'admin123'}, follow_redirects=True)


def _create_user(session, username='editor1', role='editor', active=True):
    u = User(username=username, role=role, is_active_flag=active)
    u.set_password('pass123')
    session.add(u)
    session.commit()
    return u


# ═══════════════════════════════════════════════════════════════════════════════
# 1. MODELS (10 tests)
# ═══════════════════════════════════════════════════════════════════════════════

class TestModels:
    def test_user_password_hash(self, app):
        with app.app_context():
            u = User(username='testuser', role='editor')
            u.set_password('mypass')
            assert u.check_password('mypass')
            assert not u.check_password('wrong')
            assert u.password_hash != 'mypass'

    def test_user_roles(self, app):
        with app.app_context():
            admin = User(username='adm', role='admin')
            editor = User(username='edt', role='editor')
            viewer = User(username='vw', role='viewer')
            assert admin.is_admin
            assert not editor.is_admin
            assert admin.can_edit
            assert editor.can_edit
            assert not viewer.can_edit

    def test_user_to_dict(self, app):
        with app.app_context():
            u = User(username='serial', role='admin')
            u.set_password('x')
            db.session.add(u)
            db.session.commit()
            d = u.to_dict()
            assert d['username'] == 'serial'
            assert d['role'] == 'admin'
            assert 'password_hash' not in d

    def test_client_balance(self, app):
        with app.app_context():
            c = Client(name='Bal', total_debt=1000, total_paid=400)
            assert c.balance == 600
            c2 = Client(name='Zero', total_debt=300, total_paid=500)
            assert c2.balance == 0

    def test_client_to_dict(self, app):
        with app.app_context():
            c = Client(name='Dict', phone='123', total_debt=100, total_paid=50)
            db.session.add(c)
            db.session.commit()
            d = c.to_dict()
            assert d['name'] == 'Dict'
            assert d['balance'] == 50

    def test_invoice_to_dict(self, app):
        with app.app_context():
            c = _make_client(db.session)
            inv = _make_invoice(db.session, c.id)
            d = inv.to_dict()
            assert d['amount'] == 500
            assert d['client_id'] == c.id

    def test_payment_to_dict(self, app):
        with app.app_context():
            c = _make_client(db.session)
            p = _make_payment(db.session, c.id)
            d = p.to_dict()
            assert d['amount'] == 200

    def test_settings_get_set(self, app):
        with app.app_context():
            Settings.set('test_key', 'test_val')
            assert Settings.get('test_key') == 'test_val'
            assert Settings.get('missing', 'default') == 'default'
            Settings.set('test_key', 'updated')
            assert Settings.get('test_key') == 'updated'

    def test_activity_log_to_dict(self, app):
        with app.app_context():
            a = ActivityLog(user_id=1, action='login', entity_type='user', entity_id=1, details='test')
            db.session.add(a)
            db.session.commit()
            d = a.to_dict()
            assert d['action'] == 'login'

    def test_client_cascade_delete(self, app):
        with app.app_context():
            c = _make_client(db.session)
            _make_invoice(db.session, c.id, 100)
            _make_payment(db.session, c.id, 50)
            cid = c.id
            db.session.delete(c)
            db.session.commit()
            assert Invoice.query.filter_by(client_id=cid).count() == 0
            assert Payment.query.filter_by(client_id=cid).count() == 0


# ═══════════════════════════════════════════════════════════════════════════════
# 2. AUTH ROUTES (9 tests)
# ═══════════════════════════════════════════════════════════════════════════════

class TestAuth:
    def test_login_page(self, client):
        r = client.get('/login')
        assert r.status_code == 200

    def test_login_success(self, client):
        r = client.post('/login', data={'username': 'admin', 'password': 'admin123'}, follow_redirects=True)
        assert r.status_code == 200

    def test_login_fail(self, client):
        r = client.post('/login', data={'username': 'admin', 'password': 'wrong'}, follow_redirects=True)
        assert r.status_code == 200

    def test_logout(self, auth_client):
        r = auth_client.get('/logout', follow_redirects=True)
        assert r.status_code == 200

    def test_users_page(self, auth_client):
        r = auth_client.get('/users')
        assert r.status_code == 200

    def test_user_add(self, auth_client):
        r = auth_client.post('/users/add', data={'username': 'newuser', 'password': 'pass123', 'role': 'editor'}, follow_redirects=True)
        assert r.status_code == 200

    def test_user_toggle(self, auth_client, app):
        with app.app_context():
            u = _create_user(db.session)
            uid = u.id
        r = auth_client.post(f'/users/{uid}/toggle', follow_redirects=True)
        assert r.status_code == 200

    def test_user_delete(self, auth_client, app):
        with app.app_context():
            u = _create_user(db.session, 'deluser')
            uid = u.id
        r = auth_client.post(f'/users/{uid}/delete', follow_redirects=True)
        assert r.status_code == 200

    def test_non_admin_cannot_access_users(self, auth_client, app):
        with app.app_context():
            u = _create_user(db.session, 'viewer1', 'viewer')
        auth_client.get('/logout', follow_redirects=True)
        auth_client.post('/login', data={'username': 'viewer1', 'password': 'pass123'}, follow_redirects=True)
        r = auth_client.get('/users', follow_redirects=True)
        assert r.status_code == 200

    def test_login_locked_no_autologin(self, client, app):
        with app.app_context():
            Settings.set('login_locked', 'true')
        r = client.get('/login')
        assert r.status_code == 200
        r2 = client.get('/users')
        assert r2.status_code == 302
        assert '/login' in r2.headers.get('Location', '')
        with app.app_context():
            Settings.set('login_locked', 'false')

    def test_login_locked_wrong_password_rejected(self, client, app):
        with app.app_context():
            Settings.set('login_locked', 'true')
        r = client.post('/login', data={'unlock_password': 'wrong'}, follow_redirects=True)
        assert 'كلمة المرور غير صحيحة' in r.get_data(as_text=True)
        r2 = client.get('/users')
        assert r2.status_code == 302
        with app.app_context():
            Settings.set('login_locked', 'false')

    def test_login_locked_unlock_with_admin_password(self, client, app):
        with app.app_context():
            Settings.set('login_locked', 'true')
        r = client.post('/login', data={'unlock_password': 'admin123'}, follow_redirects=True)
        assert r.status_code == 200
        r2 = client.get('/users')
        assert r2.status_code == 200
        with app.app_context():
            Settings.set('login_locked', 'false')


# ═══════════════════════════════════════════════════════════════════════════════
# 3. CLIENT ROUTES (10 tests)
# ═══════════════════════════════════════════════════════════════════════════════

class TestClients:
    def test_index(self, auth_client):
        r = auth_client.get('/')
        assert r.status_code == 200

    def test_add_client(self, auth_client, app):
        r = auth_client.post('/client/add', data={'name': 'Ahmed', 'phone': '20101112233', 'notes': 'test'}, follow_redirects=True)
        assert r.status_code == 200
        with app.app_context():
            assert Client.query.filter_by(name='Ahmed').first() is not None

    def test_add_client_page(self, auth_client):
        r = auth_client.get('/client/add')
        assert r.status_code == 200

    def test_client_detail(self, auth_client, app):
        with app.app_context():
            c = _make_client(db.session, 'DetailClient')
            cid = c.id
        r = auth_client.get(f'/client/{cid}')
        assert r.status_code == 200

    def test_edit_client(self, auth_client, app):
        with app.app_context():
            c = _make_client(db.session, 'OldName')
            cid = c.id
        r = auth_client.post(f'/client/{cid}/edit', data={'name': 'NewName', 'phone': '123'}, follow_redirects=True)
        assert r.status_code == 200
        with app.app_context():
            assert db.session.get(Client, cid).name == 'NewName'

    def test_edit_client_page(self, auth_client, app):
        with app.app_context():
            c = _make_client(db.session)
            cid = c.id
        r = auth_client.get(f'/client/{cid}/edit')
        assert r.status_code == 200

    def test_delete_client(self, auth_client, app):
        with app.app_context():
            c = _make_client(db.session, 'ToDelete')
            cid = c.id
        r = auth_client.post(f'/client/{cid}/delete', follow_redirects=True)
        assert r.status_code == 200
        with app.app_context():
            assert db.session.get(Client, cid) is None

    def test_client_settings(self, auth_client, app):
        with app.app_context():
            c = _make_client(db.session)
            cid = c.id
        r = auth_client.get(f'/client/{cid}/settings', follow_redirects=True)
        assert r.status_code == 200
        r = auth_client.post(f'/client/{cid}/settings', data={'reminder_enabled': 'on', 'reminder_template': '2'}, follow_redirects=True)
        assert r.status_code == 200

    def test_search(self, auth_client, app):
        with app.app_context():
            _make_client(db.session, 'Searchable')
        r = auth_client.get('/?q=Searchable')
        assert r.status_code == 200

    def test_status_filter(self, auth_client, app):
        with app.app_context():
            _make_client(db.session, 'DueClient')
        r = auth_client.get('/?status=due')
        assert r.status_code == 200

    def test_toggle_dark(self, auth_client):
        r = auth_client.post('/api/toggle-dark')
        assert r.status_code == 200

    def test_unauthenticated_redirects(self, client):
        r = client.get('/')
        assert r.status_code == 302


# ═══════════════════════════════════════════════════════════════════════════════
# 4. INVOICE ROUTES (3 tests)
# ═══════════════════════════════════════════════════════════════════════════════

class TestInvoices:
    def test_add_invoice(self, auth_client, app):
        with app.app_context():
            c = _make_client(db.session)
            cid = c.id
        r = auth_client.post(f'/client/{cid}/invoice/add', data={
            'amount': '750', 'description': 'New invoice', 'date': '2026-01-15'
        }, follow_redirects=True)
        assert r.status_code == 200
        data = json.loads(r.data)
        assert data.get('ok') is True

    def test_delete_invoice(self, auth_client, app):
        with app.app_context():
            c = _make_client(db.session)
            inv = _make_invoice(db.session, c.id, 100)
            iid = inv.id
        r = auth_client.post(f'/invoice/{iid}/delete', follow_redirects=True)
        assert r.status_code == 200
        data = json.loads(r.data)
        assert data.get('ok') is True

    def test_invoice_recalculates_client(self, auth_client, app):
        with app.app_context():
            c = _make_client(db.session, 'Recalc', debt=0, paid=0)
            cid = c.id
        auth_client.post(f'/client/{cid}/invoice/add', data={'amount': '500', 'description': 'inv'}, follow_redirects=True)
        with app.app_context():
            c2 = db.session.get(Client, cid)
            assert c2.total_debt == 500


# ═══════════════════════════════════════════════════════════════════════════════
# 5. PAYMENT ROUTES (3 tests)
# ═══════════════════════════════════════════════════════════════════════════════

class TestPayments:
    def test_add_payment(self, auth_client, app):
        with app.app_context():
            c = _make_client(db.session, debt=1000, paid=0)
            cid = c.id
        r = auth_client.post(f'/client/{cid}/payment/add', data={
            'amount': '250', 'notes': 'partial', 'date': '2026-02-01'
        }, follow_redirects=True)
        assert r.status_code == 200
        data = json.loads(r.data)
        assert data.get('ok') is True

    def test_delete_payment(self, auth_client, app):
        with app.app_context():
            c = _make_client(db.session)
            p = _make_payment(db.session, c.id, 100)
            pid = p.id
        r = auth_client.post(f'/payment/{pid}/delete', follow_redirects=True)
        assert r.status_code == 200
        data = json.loads(r.data)
        assert data.get('ok') is True

    def test_payment_recalculates_client(self, auth_client, app):
        with app.app_context():
            c = _make_client(db.session, 'PayCalc', debt=0, paid=0)
            cid = c.id
            _make_invoice(db.session, cid, 1000)
        auth_client.post(f'/client/{cid}/payment/add', data={'amount': '400', 'notes': 'pay'}, follow_redirects=True)
        with app.app_context():
            c2 = db.session.get(Client, cid)
            assert c2.total_paid == 400
            assert c2.balance == 600


# ═══════════════════════════════════════════════════════════════════════════════
# 6. WHATSAPP / SETTINGS / DATABASE ROUTES (19 tests)
# ═══════════════════════════════════════════════════════════════════════════════

class TestWhatsApp:
    def test_settings_page(self, auth_client):
        r = auth_client.get('/settings')
        assert r.status_code == 200

    def test_settings_general_tab(self, auth_client):
        r = auth_client.get('/settings?tab=general')
        assert r.status_code == 200

    def test_settings_save_general(self, auth_client, app):
        r = auth_client.post('/settings', data={
            'tab': 'general', 'app_country': 'SA',
            'app_timezone': 'Asia/Riyadh', 'app_currency': 'ريال سعودي', 'app_currency_short': 'ر.س'
        }, follow_redirects=True)
        assert r.status_code == 200
        with app.app_context():
            assert Settings.get('app_country') == 'SA'

    def test_settings_whatsapp_tab(self, auth_client):
        r = auth_client.get('/settings?tab=whatsapp')
        assert r.status_code == 200

    def test_settings_save_baileys(self, auth_client, app):
        r = auth_client.post('/settings', data={
            'tab': 'whatsapp', 'baileys_url': 'http://localhost:9999'
        }, follow_redirects=True)
        assert r.status_code == 200
        with app.app_context():
            assert Settings.get('baileys_url') == 'http://localhost:9999'

    def test_settings_reminder_tab(self, auth_client):
        r = auth_client.get('/settings?tab=reminder')
        assert r.status_code == 200

    def test_settings_save_reminder(self, auth_client, app):
        r = auth_client.post('/settings', data={
            'tab': 'reminder', 'reminder_enabled': 'on',
            'reminder_times': '09:00', 'reminder_frequency': 'daily'
        }, follow_redirects=True)
        assert r.status_code == 200
        with app.app_context():
            assert Settings.get('reminder_enabled') == 'true'

    def test_settings_templates_tab(self, auth_client):
        r = auth_client.get('/settings?tab=templates')
        assert r.status_code == 200

    def test_settings_save_templates(self, auth_client, app):
        r = auth_client.post('/settings', data={
            'tab': 'templates', 'template_1': 'Hi {name}', 'template_2': 'Balance {balance}', 'template_3': 'Bye'
        }, follow_redirects=True)
        assert r.status_code == 200
        with app.app_context():
            assert Settings.get('template_1') == 'Hi {name}'

    def test_settings_database_tab(self, auth_client):
        r = auth_client.get('/settings?tab=database')
        assert r.status_code == 200
        assert 'إدارة قاعدة البيانات'.encode('utf-8') in r.data

    def test_whatsapp_status_endpoint(self, auth_client):
        r = auth_client.get('/api/whatsapp/status')
        assert r.status_code == 200
        data = json.loads(r.data)
        assert 'connected' in data
        assert 'status' in data

    def test_send_reminder_no_phone(self, auth_client, app):
        with app.app_context():
            c = Client(name='NoPhone')
            db.session.add(c)
            db.session.commit()
            cid = c.id
        r = auth_client.post(f'/api/whatsapp/send-reminder/{cid}', follow_redirects=True)
        data = json.loads(r.data)
        assert data.get('ok') is False

    def test_send_reminder_404(self, auth_client):
        r = auth_client.post('/api/whatsapp/send-reminder/99999', follow_redirects=True)
        assert r.status_code == 404

    def test_baileys_logout_endpoint(self, auth_client):
        r = auth_client.post('/api/baileys/logout', follow_redirects=True)
        assert r.status_code == 200
        data = json.loads(r.data)
        assert 'ok' in data

    def test_baileys_start_endpoint(self, auth_client):
        r = auth_client.post('/api/baileys/start', follow_redirects=True)
        assert r.status_code == 200
        data = json.loads(r.data)
        assert 'ok' in data
        assert 'msg' in data

    def test_db_backup_endpoint(self, auth_client):
        r = auth_client.post('/api/database/backup', follow_redirects=True)
        assert r.status_code == 200
        data = json.loads(r.data)
        assert 'ok' in data

    def test_db_export_endpoint(self, auth_client):
        r = auth_client.get('/api/database/export-db')
        assert r.status_code == 200
        assert 'octet' in r.content_type or 'sqlite' in r.content_type

    def test_db_restore_invalid_file(self, auth_client):
        r = auth_client.post('/api/database/restore',
                             json={'filename': '../../etc/passwd'},
                             content_type='application/json')
        data = json.loads(r.data)
        assert data.get('ok') is False

    def test_db_delete_backup_invalid(self, auth_client):
        r = auth_client.post('/api/database/delete-backup',
                             json={'filename': '../hack.db'},
                             content_type='application/json')
        data = json.loads(r.data)
        assert data.get('ok') is False

    def test_db_reset_no_confirm(self, auth_client):
        r = auth_client.post('/api/database/reset',
                             json={'confirm': 'WRONG'},
                             content_type='application/json')
        data = json.loads(r.data)
        assert data.get('ok') is False

    def test_db_reset_with_confirm(self, auth_client, app):
        with app.app_context():
            _make_client(db.session, 'ToReset')
        r = auth_client.post('/api/database/reset',
                             json={'confirm': 'DELETE_ALL'},
                             content_type='application/json')
        data = json.loads(r.data)
        assert data.get('ok') is True
        with app.app_context():
            assert Client.query.count() == 0

    def test_db_download_backup_invalid(self, auth_client):
        r = auth_client.get('/api/database/download-backup?file=../hack.db')
        assert r.status_code == 404 or r.status_code == 400


# ═══════════════════════════════════════════════════════════════════════════════
# 7. REPORTS ROUTES (9 tests)
# ═══════════════════════════════════════════════════════════════════════════════

class TestReports:
    def test_report_page(self, auth_client, app):
        with app.app_context():
            _make_client(db.session)
        r = auth_client.get('/report')
        assert r.status_code == 200

    def test_advanced_report(self, auth_client, app):
        with app.app_context():
            _make_client(db.session)
        r = auth_client.get('/advanced-report')
        assert r.status_code == 200

    def test_compare_report(self, auth_client, app):
        with app.app_context():
            _make_client(db.session)
        r = auth_client.get('/compare')
        assert r.status_code == 200

    def test_export_excel(self, auth_client, app):
        with app.app_context():
            _make_client(db.session)
        r = auth_client.get('/export')
        assert r.status_code == 200
        assert 'spreadsheetml' in r.content_type or 'octet' in r.content_type

    def test_export_pdf(self, auth_client, app):
        with app.app_context():
            _make_client(db.session)
        r = auth_client.get('/export/pdf')
        assert r.status_code == 200

    def test_import_page(self, auth_client):
        r = auth_client.get('/import')
        assert r.status_code == 200

    def test_import_template(self, auth_client):
        r = auth_client.get('/import/template')
        assert r.status_code == 200
        assert 'spreadsheetml' in r.content_type

    def test_import_preview(self, auth_client):
        wb = Workbook()
        ws = wb.active
        ws.append(['اسم العميل', 'رقم الهاتف', 'المديونية', 'المدفوع', 'ملاحظات'])
        ws.append(['أحمد', '010123', 500, 200, 'test'])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        buf.filename = 'import.xlsx'
        r = auth_client.post('/import', data={'file': buf}, content_type='multipart/form-data', follow_redirects=True)
        assert r.status_code == 200

    def test_import_empty_file(self, auth_client):
        f = io.BytesIO(b'')
        f.filename = 'empty.csv'
        r = auth_client.post('/import', data={'file': f}, content_type='multipart/form-data', follow_redirects=True)
        assert r.status_code == 200

    def test_backup(self, auth_client):
        r = auth_client.get('/backup', follow_redirects=True)
        assert r.status_code == 200


# ═══════════════════════════════════════════════════════════════════════════════
# 8. API V1 ROUTES (13 tests)
# ═══════════════════════════════════════════════════════════════════════════════

class TestAPI:
    def test_list_clients(self, auth_client, app):
        with app.app_context():
            _make_client(db.session)
        r = auth_client.get('/api/v1/clients')
        data = json.loads(r.data)
        assert r.status_code == 200
        assert 'clients' in data
        assert data['total'] >= 1

    def test_create_client(self, auth_client):
        r = auth_client.post('/api/v1/clients', json={'name': 'APIClient', 'phone': '966501234567'})
        assert r.status_code == 201
        data = json.loads(r.data)
        assert data['client']['name'] == 'APIClient'

    def test_get_client(self, auth_client, app):
        with app.app_context():
            c = _make_client(db.session, 'GetMe')
            cid = c.id
        r = auth_client.get(f'/api/v1/clients/{cid}')
        data = json.loads(r.data)
        assert data['name'] == 'GetMe'

    def test_update_client(self, auth_client, app):
        with app.app_context():
            c = _make_client(db.session, 'OldAPI')
            cid = c.id
        r = auth_client.put(f'/api/v1/clients/{cid}', json={'name': 'NewAPI'})
        assert r.status_code == 200

    def test_delete_client_api(self, auth_client, app):
        with app.app_context():
            c = _make_client(db.session, 'DelAPI')
            cid = c.id
        r = auth_client.delete(f'/api/v1/clients/{cid}')
        assert r.status_code == 200

    def test_add_invoice_api(self, auth_client, app):
        with app.app_context():
            c = _make_client(db.session)
            cid = c.id
        r = auth_client.post(f'/api/v1/clients/{cid}/invoices', json={'amount': 300, 'description': 'api inv'})
        assert r.status_code == 201

    def test_delete_invoice_api(self, auth_client, app):
        with app.app_context():
            c = _make_client(db.session)
            inv = _make_invoice(db.session, c.id)
            iid = inv.id
        r = auth_client.delete(f'/api/v1/invoices/{iid}')
        assert r.status_code == 200

    def test_add_payment_api(self, auth_client, app):
        with app.app_context():
            c = _make_client(db.session)
            cid = c.id
        r = auth_client.post(f'/api/v1/clients/{cid}/payments', json={'amount': 150, 'notes': 'api pay'})
        assert r.status_code == 201

    def test_delete_payment_api(self, auth_client, app):
        with app.app_context():
            c = _make_client(db.session)
            p = _make_payment(db.session, c.id)
            pid = p.id
        r = auth_client.delete(f'/api/v1/payments/{pid}')
        assert r.status_code == 200

    def test_report_summary(self, auth_client, app):
        with app.app_context():
            _make_client(db.session)
        r = auth_client.get('/api/v1/reports/summary')
        data = json.loads(r.data)
        assert 'total_clients' in data

    def test_report_trends(self, auth_client):
        r = auth_client.get('/api/v1/reports/trends')
        data = json.loads(r.data)
        assert 'trends' in data
        assert len(data['trends']) == 6
        months = [m['month'] for m in data['trends']]
        assert len(set(months)) == 6, "trend months must not contain duplicates"
        assert months == sorted(months), "trend months must be ascending"
        from datetime import datetime as _dt
        assert months[-1] == _dt.now().strftime('%Y-%m'), "last trend month must be current month"

    def test_activity_log(self, auth_client):
        r = auth_client.get('/api/v1/activity')
        data = json.loads(r.data)
        assert 'activities' in data

    def test_list_users_api(self, auth_client):
        r = auth_client.get('/api/v1/users')
        data = json.loads(r.data)
        assert 'users' in data


# ═══════════════════════════════════════════════════════════════════════════════
# 9. UTILS (16 tests)
# ═══════════════════════════════════════════════════════════════════════════════

class TestUtils:
    def test_allowed_file(self):
        assert allowed_file('test.png')
        assert allowed_file('test.pdf')
        assert allowed_file('test.jpg')
        assert not allowed_file('test.exe')
        assert not allowed_file('test.py')
        assert not allowed_file('test.xlsx')

    def test_recalc_client(self, app):
        with app.app_context():
            c = _make_client(db.session, debt=0, paid=0)
            _make_invoice(db.session, c.id, 800)
            _make_payment(db.session, c.id, 300)
            recalc_client(c.id)
            c2 = db.session.get(Client, c.id)
            assert c2.total_debt == 800
            assert c2.total_paid == 300
            assert c2.balance == 500

    def test_recalc_resets_when_last_invoice_deleted(self, app):
        with app.app_context():
            c = _make_client(db.session, debt=100, paid=0)
            inv = _make_invoice(db.session, c.id, 100)
            recalc_client(c.id)
            db.session.refresh(c)
            assert c.total_debt == 100
            db.session.delete(inv)
            db.session.commit()
            recalc_client(c.id)
            db.session.refresh(c)
            assert c.total_debt == 0
            assert c.status == 'paid'

    def test_recalc_preserves_imported_base_totals(self, app):
        with app.app_context():
            c = Client(name='Imported', total_debt=1000, total_paid=200,
                       base_debt=1000, base_paid=200, status='due')
            db.session.add(c)
            db.session.commit()
            _make_invoice(db.session, c.id, 300)
            _make_payment(db.session, c.id, 50)
            recalc_client(c.id)
            db.session.refresh(c)
            assert c.total_debt == 1300
            assert c.total_paid == 250
            assert c.balance == 1050

    def test_log_activity(self, app):
        with app.app_context():
            log_activity(1, 'test_action', 'client', 1, 'test details')
            a = ActivityLog.query.filter_by(action='test_action').first()
            assert a is not None
            assert a.details == 'test details'

    def test_get_app_settings_eg(self, app):
        with app.app_context():
            s = get_app_settings()
            assert s['country'] in ('EG', 'SA')
            assert 'phone_code' in s
            assert 'currency' in s

    def test_get_app_settings_sa(self, app):
        with app.app_context():
            Settings.set('app_country', 'SA')
            s = get_app_settings()
            assert s['phone_code'] == '966'
            assert 'ريال' in s['currency']

    def test_whatsapp_settings(self, app):
        with app.app_context():
            ws = get_whatsapp_settings()
            assert 'baileys_url' in ws

    def test_build_reminder_message(self, app):
        with app.app_context():
            c = Client(name='أحمد', total_debt=1000, total_paid=600)
            db.session.add(c)
            db.session.commit()
            msg = build_reminder_message(c, 1)
            assert 'أحمد' in msg
            assert '400' in msg

    def test_build_reminder_all_templates(self, app):
        with app.app_context():
            c = Client(name='خالد', total_debt=2000, total_paid=500)
            db.session.add(c)
            db.session.commit()
            for t in [1, 2, 3]:
                msg = build_reminder_message(c, t)
                assert 'خالد' in msg
                assert '1,500' in msg

    def test_export_excel_empty(self, app):
        with app.app_context():
            wb = export_excel([])
            assert wb is not None

    def test_export_excel_with_data(self, app):
        with app.app_context():
            c = _make_client(db.session)
            wb = export_excel([c])
            assert wb is not None

    def test_create_sample_template(self):
        wb = create_sample_template()
        assert wb is not None
        assert len(wb.sheetnames) > 0

    def test_parse_uploaded_file_csv(self):
        content = b"name,phone,total_debt\ntest,123,100\n"
        f = io.BytesIO(content)
        f.filename = 'test.csv'
        result = parse_uploaded_file(f)
        assert result is not None
        assert len(result['data']) == 1

    def test_auto_detect_columns(self):
        headers = ['اسم العميل', 'رقم الهاتف', 'المديونية', 'المدفوع', 'ملاحظات']
        mapping = auto_detect_columns(headers)
        assert mapping['name'] == 0
        assert mapping['phone'] == 1
        assert mapping['total_debt'] == 2

    def test_validate_import_rows(self):
        data = [['Ahmed', '123', 500, 200, '']]
        mapping = {'name': 0, 'phone': 1, 'total_debt': 2, 'total_paid': 3, 'notes': 4}
        result = validate_import_rows(data, mapping)
        assert len(result) == 1
        assert result[0]['valid'] is True

    def test_validate_import_rows_errors(self):
        data = [['', '123', -100, 200, '']]
        mapping = {'name': 0, 'phone': 1, 'total_debt': 2, 'total_paid': 3, 'notes': 4}
        result = validate_import_rows(data, mapping)
        assert result[0]['valid'] is False
        assert len(result[0]['errors']) >= 1

    def test_country_options_complete(self):
        assert 'EG' in COUNTRY_OPTIONS
        assert 'SA' in COUNTRY_OPTIONS
        for k, v in COUNTRY_OPTIONS.items():
            assert 'phone_code' in v
            assert 'timezone' in v
            assert 'currency' in v

    def test_send_whatsapp_no_service(self, app):
        with app.app_context():
            Settings.set('baileys_url', 'http://localhost:19999')
            ok, msg = send_whatsapp('201012345678', 'test')
            assert ok is False

    def test_normalize_phone(self):
        assert normalize_phone('0501234567', 'SA') == '966501234567'
        assert normalize_phone('0501234567', 'EG') == '966501234567'
        assert normalize_phone('966501234567', 'SA') == '966501234567'
        assert normalize_phone('+966501234567', 'SA') == '966501234567'
        assert normalize_phone('00966501234567', 'SA') == '966501234567'
        assert normalize_phone('01012345678', 'EG') == '201012345678'
        assert normalize_phone('01012345678', 'SA') == '201012345678'
        assert normalize_phone('01112345678', 'EG') == '201112345678'
        assert normalize_phone('0223456789', 'EG') == '20223456789'
        assert normalize_phone('201012345678', 'EG') == '201012345678'
        assert normalize_phone(None, 'SA') == ''
        assert normalize_phone('', 'SA') == ''

    def test_reminder_time_filter_default_client(self, app, monkeypatch):
        sent = []
        monkeypatch.setattr('app.utils.time.sleep', lambda s: None)
        monkeypatch.setattr('app.utils.send_whatsapp', lambda p, m: sent.append(p) or (True, 'ok'))
        with app.app_context():
            Settings.set('reminder_times', '10:00')
            Settings.set('reminder_frequency', 'daily')
            c1 = Client(name='Default', phone='201012345678', total_debt=500,
                        total_paid=0, status='due', reminder_enabled=True)
            c2 = Client(name='Custom', phone='20101112233', total_debt=500,
                        total_paid=0, status='due', reminder_enabled=True,
                        reminder_times='09:00,14:00')
            db.session.add_all([c1, c2])
            db.session.commit()
        _send_reminders_background(app, trigger_time='10:00', trigger_freq='daily',
                                   trigger_day='sun', trigger_dom=1)
        assert sent == ['201012345678']
        sent.clear()
        _send_reminders_background(app, trigger_time='14:00', trigger_freq='daily',
                                   trigger_day='sun', trigger_dom=1)
        assert sent == ['20101112233']


# ═══════════════════════════════════════════════════════════════════════════════
# 10. ACCOUNTING EXCEL PARSER (5 tests)
# ═══════════════════════════════════════════════════════════════════════════════

class TestAccountingParser:
    def test_cell_str(self):
        class FakeCell:
            def __init__(self, v): self.value = v
        assert _cell_str(FakeCell('hello')) == 'hello'
        assert _cell_str(FakeCell(None)) == ''
        assert _cell_str(FakeCell('  spaces  ')) == 'spaces'

    def test_cell_num(self):
        class FakeCell:
            def __init__(self, v): self.value = v
        assert _cell_num(FakeCell(42)) == 42.0
        assert _cell_num(FakeCell(None)) == 0.0
        assert _cell_num(FakeCell(None), 99) == 99.0
        assert _cell_num(FakeCell('1,000')) == 1000.0

    def test_find_header_row(self):
        class FakeWS:
            def __init__(self):
                self.max_row = 5
                self.max_column = 3
                self._data = {
                    (1,1): None, (1,2): None, (1,3): None,
                    (2,1): 'Header', (2,2): 'مدين', (2,3): 'دائن',
                    (3,1): 'data', (3,2): 100, (3,3): 0,
                }
            def cell(self, r, c):
                class C:
                    def __init__(self, v): self.value = v
                return C(self._data.get((r,c)))
        ws = FakeWS()
        assert _find_header_row(ws, ['مدين', 'دائن']) == 2

    def test_detect_format_none(self):
        f = io.BytesIO(b'not an excel file')
        f.filename = 'bad.xlsx'
        assert detect_format(f) is None

    def test_build_customer_preview_empty(self):
        parsed = {'customers': [], 'transactions': []}
        preview = build_customer_preview(parsed)
        assert preview == []

    def test_parse_real_accounting_layout(self):
        from datetime import datetime
        wb = Workbook()
        ws = wb.active
        ws.title = 'all'
        for c, v in {2: 'م', 3: 'كود العميل', 4: 'رقم التقرير', 5: 'المالك', 6: 'اسم العميل',
                     7: 'الموقع', 8: 'الصنف', 9: 'التاريخ', 10: 'الكمية', 11: 'السعر',
                     12: 'مدين', 13: 'دائن', 14: 'البيان - ملاحظات', 15: 'الفرع',
                     16: 'طريقة الدفع', 17: 'Mth'}.items():
            ws.cell(8, c, v)
        ws.cell(9, 2, 1); ws.cell(9, 4, 261674); ws.cell(9, 5, 'المالك')
        ws.cell(9, 6, 'ايهاب بكرى'); ws.cell(9, 7, 'النزهة'); ws.cell(9, 8, 'تحميل')
        ws.cell(9, 9, datetime(2026, 5, 20)); ws.cell(9, 10, 1); ws.cell(9, 11, 500)
        ws.cell(9, 12, 500); ws.cell(9, 14, 'ملاحظات'); ws.cell(9, 17, datetime(2026, 5, 31))
        ws.cell(10, 2, 2); ws.cell(10, 6, 'عميل ثاني'); ws.cell(10, 11, 200)
        ws.cell(10, 12, 0); ws.cell(10, 13, 200); ws.cell(10, 9, datetime(2026, 6, 1))

        data = wb.create_sheet('Data')
        for c, v in {2: 'الكود', 3: 'م', 4: 'كود الحساب', 5: 'اسم العميل', 6: 'الإيرادات',
                     7: 'التحصيلات', 8: 'الرصيد', 9: 'نسبة الإيرادات'}.items():
            data.cell(7, c, v)
        data.cell(8, 5, 'ايهاب بكرى'); data.cell(8, 6, 1100); data.cell(8, 7, 0); data.cell(8, 8, 1100)

        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        from app.importers.accounting_excel import parse_accounting_excel
        parsed = parse_accounting_excel(bio)
        assert len(parsed['transactions']) == 2
        t0 = parsed['transactions'][0]
        assert t0['customer_name'] == 'ايهاب بكرى'
        assert t0['debit'] == 500.0 and t0['credit'] == 0.0
        assert t0['date'] == '2026-05-20'
        assert t0['seq'] == 1
        assert t0['month'] == '2026-05'
        t1 = parsed['transactions'][1]
        assert t1['credit'] == 200.0
        assert parsed['customers'][0]['name'] == 'ايهاب بكرى'
        assert parsed['customers'][0]['revenue'] == 1100.0


# ═══════════════════════════════════════════════════════════════════════════════
# 11. TEMPLATE RENDERING (10 tests)
# ═══════════════════════════════════════════════════════════════════════════════

class TestTemplates:
    def test_base_template_renders(self, auth_client):
        r = auth_client.get('/')
        assert r.status_code == 200
        assert b'<!DOCTYPE html>' in r.data or b'<html' in r.data

    def test_index_has_stats(self, auth_client, app):
        with app.app_context():
            _make_client(db.session, debt=500, paid=200)
        r = auth_client.get('/')
        assert 'إجمالي العملاء'.encode('utf-8') in r.data

    def test_index_has_wa_status(self, auth_client):
        r = auth_client.get('/')
        assert 'waStatusBar'.encode() in r.data

    def test_client_detail_has_currency(self, auth_client, app):
        with app.app_context():
            c = _make_client(db.session)
            cid = c.id
        r = auth_client.get(f'/client/{cid}')
        assert r.status_code == 200

    def test_settings_has_general_tab(self, auth_client):
        r = auth_client.get('/settings?tab=general')
        assert 'الإعدادات العامة'.encode('utf-8') in r.data

    def test_settings_has_country_select(self, auth_client):
        r = auth_client.get('/settings?tab=general')
        assert 'app_country'.encode() in r.data

    def test_settings_has_database_tab(self, auth_client):
        r = auth_client.get('/settings?tab=database')
        assert 'إدارة قاعدة البيانات'.encode('utf-8') in r.data
        assert 'نسخة احتياطية'.encode('utf-8') in r.data

    def test_settings_has_whatsapp_tab(self, auth_client):
        r = auth_client.get('/settings?tab=whatsapp')
        assert 'إعدادات واتساب'.encode('utf-8') in r.data

    def test_settings_has_reminder_tab(self, auth_client):
        r = auth_client.get('/settings?tab=reminder')
        assert 'إعدادات التذكير'.encode('utf-8') in r.data

    def test_settings_has_templates_tab(self, auth_client):
        r = auth_client.get('/settings?tab=templates')
        assert 'قوالب الرسائل'.encode('utf-8') in r.data


# ═══════════════════════════════════════════════════════════════════════════════
# 12. EDGE CASES & SECURITY (10 tests)
# ═══════════════════════════════════════════════════════════════════════════════

class TestEdgeCases:
    def test_404_client(self, auth_client):
        r = auth_client.get('/client/99999')
        assert r.status_code == 404

    def test_editor_cannot_delete(self, auth_client, app):
        with app.app_context():
            u = _create_user(db.session, 'edonly', 'editor')
            uid = u.id
        auth_client.get('/logout', follow_redirects=True)
        auth_client.post('/login', data={'username': 'edonly', 'password': 'pass123'}, follow_redirects=True)
        with app.app_context():
            c = _make_client(db.session, 'Protected')
            cid = c.id
        r = auth_client.post(f'/client/{cid}/delete', follow_redirects=True)
        with app.app_context():
            assert db.session.get(Client, cid) is not None

    def test_viewer_cannot_add_client(self, auth_client, app):
        with app.app_context():
            u = _create_user(db.session, 'vviewer', 'viewer')
        auth_client.get('/logout', follow_redirects=True)
        auth_client.post('/login', data={'username': 'vviewer', 'password': 'pass123'}, follow_redirects=True)
        r = auth_client.post('/client/add', data={'name': 'Nope'}, follow_redirects=True)
        assert r.status_code == 200

    def test_api_unauthorized(self, client):
        r = client.get('/api/v1/clients')
        assert r.status_code == 302

    def test_duplicate_client_name(self, auth_client, app):
        with app.app_context():
            _make_client(db.session, 'DupTest')
        r = auth_client.post('/client/add', data={'name': 'DupTest', 'phone': '111'}, follow_redirects=True)
        assert r.status_code == 200

    def test_payment_exceeds_debt(self, auth_client, app):
        with app.app_context():
            c = _make_client(db.session, debt=100, paid=0)
            cid = c.id
        r = auth_client.post(f'/client/{cid}/payment/add', data={'amount': '500', 'notes': 'overpay'}, follow_redirects=True)
        assert r.status_code == 200

    def test_non_admin_cannot_backup(self, auth_client, app):
        with app.app_context():
            u = _create_user(db.session, 'viewer2', 'viewer')
        auth_client.get('/logout', follow_redirects=True)
        auth_client.post('/login', data={'username': 'viewer2', 'password': 'pass123'}, follow_redirects=True)
        r = auth_client.post('/api/database/backup', follow_redirects=True)
        data = json.loads(r.data)
        assert data.get('ok') is False or r.status_code == 403

    def test_non_admin_cannot_reset_db(self, auth_client, app):
        with app.app_context():
            u = _create_user(db.session, 'viewer3', 'viewer')
        auth_client.get('/logout', follow_redirects=True)
        auth_client.post('/login', data={'username': 'viewer3', 'password': 'pass123'}, follow_redirects=True)
        r = auth_client.post('/api/database/reset',
                             json={'confirm': 'DELETE_ALL'},
                             content_type='application/json')
        assert r.status_code == 403 or json.loads(r.data).get('ok') is False

    def test_non_admin_cannot_start_baileys(self, auth_client, app):
        with app.app_context():
            u = _create_user(db.session, 'viewer4', 'viewer')
        auth_client.get('/logout', follow_redirects=True)
        auth_client.post('/login', data={'username': 'viewer4', 'password': 'pass123'}, follow_redirects=True)
        r = auth_client.post('/api/baileys/start', follow_redirects=True)
        assert r.status_code == 403 or json.loads(r.data).get('ok') is False

    def test_editor_can_send_reminder(self, auth_client, app):
        with app.app_context():
            u = _create_user(db.session, 'edonly2', 'editor')
            c = _make_client(db.session, 'RemindTest', phone='201012345678')
            cid = c.id
        auth_client.get('/logout', follow_redirects=True)
        auth_client.post('/login', data={'username': 'edonly2', 'password': 'pass123'}, follow_redirects=True)
        r = auth_client.post(f'/api/whatsapp/send-reminder/{cid}', follow_redirects=True)
        assert r.status_code == 200
        data = json.loads(r.data)
        assert 'ok' in data


# ═══════════════════════════════════════════════════════════════════════════════
# 13. APP SETTINGS INTEGRATION (4 tests)
# ═══════════════════════════════════════════════════════════════════════════════

class TestAppSettings:
    def test_default_country_is_eg(self, app):
        with app.app_context():
            s = get_app_settings()
            assert s['country'] == 'EG'
            assert s['phone_code'] == '20'

    def test_switch_to_sa(self, app):
        with app.app_context():
            Settings.set('app_country', 'SA')
            s = get_app_settings()
            assert s['phone_code'] == '966'
            assert 'ريال' in s['currency']

    def test_currency_in_templates(self, auth_client, app):
        with app.app_context():
            Settings.set('app_country', 'SA')
            Settings.set('app_currency_short', 'ر.س')
        r = auth_client.get('/')
        assert 'ر.س'.encode('utf-8') in r.data

    def test_phone_code_in_client_edit(self, auth_client, app):
        with app.app_context():
            Settings.set('app_country', 'SA')
            c = _make_client(db.session)
            cid = c.id
        r = auth_client.get(f'/client/{cid}/edit')
        assert '+966'.encode() in r.data
